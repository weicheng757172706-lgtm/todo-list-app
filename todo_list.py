#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Todo List Application with Excel-like interface and priority matrix
A task management application with Eisenhower Matrix prioritization
"""

import json
import os
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import shutil


class TodoApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Todo List - 每日任务管理")
        self.root.geometry("1000x700")
        
        # Configure style for Excel-like appearance
        self.setup_styles()
        
        # Data file path
        self.data_file = os.path.join(os.path.expanduser("~"), "Documents", "todo_data.json")
        os.makedirs(os.path.dirname(self.data_file), exist_ok=True)
        
        # Priority options (Eisenhower Matrix)
        self.priority_options = ["重要紧急", "重要不紧急", "不重要紧急", "不重要不紧急"]
        self.priority_colors = {
            "重要紧急": "#FF6B6B",      # Red
            "重要不紧急": "#4ECDC4",    # Teal
            "不重要紧急": "#FFE66D",    # Yellow
            "不重要不紧急": "#95E1D3"  # Light green
        }
        
        # Load data
        self.data = self.load_data()
        
        # Setup GUI
        self.setup_gui()
        
        # Refresh display
        self.refresh_display()
    
    def setup_styles(self):
        """Setup ttk styles for Excel-like appearance"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Configure Treeview for Excel-like appearance
        style.configure("Treeview",
                      background="white",
                      foreground="black",
                      rowheight=25,
                      fieldbackground="white",
                      borderwidth=1,
                      relief="solid")
        
        style.configure("Treeview.Heading",
                      background="#4F81BD",
                      foreground="white",
                      font=("Microsoft YaHei", 9, "bold"),
                      relief="raised",
                      borderwidth=1)
        
        style.map("Treeview",
                 background=[('selected', '#CCE5FF')],
                 foreground=[('selected', 'black')])
    
    def load_data(self):
        """Load data from JSON file"""
        if os.path.exists(self.data_file):
            try:
                with open(self.data_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                pass
        return {"tasks": [], "completed": []}
    
    def save_data(self):
        """Save data to JSON file"""
        with open(self.data_file, 'w', encoding='utf-8') as f:
            json.dump(self.data, f, ensure_ascii=False, indent=2)
    
    def setup_gui(self):
        """Setup the GUI components"""
        # Title
        title_frame = ttk.Frame(self.root, padding="10")
        title_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E))
        
        title_label = ttk.Label(title_frame, text="📝 Todo List - 每日任务管理", 
                               font=("Microsoft YaHei", 16, "bold"))
        title_label.pack()
        
        # Input frame
        input_frame = ttk.LabelFrame(self.root, text="添加新任务", padding="10")
        input_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), padx=10, pady=5)
        
        ttk.Label(input_frame, text="任务内容:").grid(row=0, column=0, padx=(0, 5), sticky=tk.W)
        self.task_entry = ttk.Entry(input_frame, width=40, font=("Microsoft YaHei", 10))
        self.task_entry.grid(row=0, column=1, padx=(0, 10))
        self.task_entry.bind('<Return>', lambda e: self.add_task())
        
        ttk.Label(input_frame, text="重要紧急程度:").grid(row=0, column=2, padx=(10, 5), sticky=tk.W)
        self.priority_var = tk.StringVar(value=self.priority_options[0])
        self.priority_combo = ttk.Combobox(input_frame, textvariable=self.priority_var,
                                           values=self.priority_options, state='readonly',
                                           width=15, font=("Microsoft YaHei", 9))
        self.priority_combo.grid(row=0, column=3, padx=(0, 10))
        
        add_btn = ttk.Button(input_frame, text="✚ 添加任务", command=self.add_task)
        add_btn.grid(row=0, column=4, padx=(10, 0))
        
        # Notebook for tabs
        notebook = ttk.Notebook(self.root)
        notebook.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), 
                     padx=10, pady=10)
        
        # Pending tasks tab
        pending_frame = ttk.Frame(notebook, padding="10")
        notebook.add(pending_frame, text="📋 待完成任务")
        self.setup_pending_tab(pending_frame)
        
        # Completed tasks tab
        completed_frame = ttk.Frame(notebook, padding="10")
        notebook.add(completed_frame, text="✅ 已完成任务")
        self.setup_completed_tab(completed_frame)
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(2, weight=1)
    
    def setup_pending_tab(self, parent):
        """Setup pending tasks tab with Excel-like table"""
        # Toolbar
        toolbar = ttk.Frame(parent)
        toolbar.grid(row=0, column=0, columnspan=4, sticky=(tk.W, tk.E), pady=(0, 10))
        
        complete_btn = ttk.Button(toolbar, text="✓ 标记完成", command=self.complete_task)
        complete_btn.grid(row=0, column=0, padx=(0, 5))
        
        delete_btn = ttk.Button(toolbar, text="🗑 删除任务", command=self.delete_task)
        delete_btn.grid(row=0, column=1, padx=5)
        
        export_btn = ttk.Button(toolbar, text="📊 导出Excel", command=self.export_to_excel)
        export_btn.grid(row=0, column=2, padx=5)
        
        refresh_btn = ttk.Button(toolbar, text="🔄 刷新", command=self.refresh_display)
        refresh_btn.grid(row=0, column=3, padx=5)
        
        # Excel-like table (Treeview)
        table_frame = ttk.Frame(parent)
        table_frame.grid(row=1, column=0, columnspan=4, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        columns = ('ID', '任务内容', '重要紧急程度', '创建时间', '状态')
        self.pending_tree = ttk.Treeview(table_frame, columns=columns, show='headings', 
                                        height=20, selectmode='browse')
        
        # Configure columns
        col_widths = {'ID': 50, '任务内容': 400, '重要紧急程度': 120, '创建时间': 150, '状态': 80}
        for col in columns:
            self.pending_tree.heading(col, text=col, anchor=tk.CENTER)
            self.pending_tree.column(col, width=col_widths.get(col, 100), anchor=tk.CENTER)
        
        self.pending_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.pending_tree.yview)
        v_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.pending_tree.configure(yscrollcommand=v_scrollbar.set)
        
        h_scrollbar = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.pending_tree.xview)
        h_scrollbar.grid(row=1, column=0, sticky=(tk.W, tk.E))
        self.pending_tree.configure(xscrollcommand=h_scrollbar.set)
        
        # Configure grid weights
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(1, weight=1)
    
    def setup_completed_tab(self, parent):
        """Setup completed tasks tab"""
        # Filter frame
        filter_frame = ttk.Frame(parent)
        filter_frame.grid(row=0, column=0, columnspan=4, sticky=(tk.W, tk.E), pady=(0, 10))
        
        ttk.Label(filter_frame, text="筛选日期:").grid(row=0, column=0, padx=(0, 10))
        
        self.date_filter = ttk.Combobox(filter_frame, width=20, state='readonly')
        self.date_filter.grid(row=0, column=1, padx=(0, 10))
        self.date_filter.bind('<<ComboboxSelected>>', lambda e: self.filter_completed())
        
        # Toolbar
        toolbar = ttk.Frame(parent)
        toolbar.grid(row=1, column=0, columnspan=4, sticky=(tk.W, tk.E), pady=(0, 10))
        
        delete_completed_btn = ttk.Button(toolbar, text="🗑 删除记录", command=self.delete_completed)
        delete_completed_btn.grid(row=0, column=0, padx=(0, 5))
        
        export_completed_btn = ttk.Button(toolbar, text="📊 导出Excel", 
                                         command=lambda: self.export_to_excel(completed=True))
        export_completed_btn.grid(row=0, column=1, padx=5)
        
        refresh_completed_btn = ttk.Button(toolbar, text="🔄 刷新", 
                                          command=self.refresh_completed)
        refresh_completed_btn.grid(row=0, column=2, padx=5)
        
        # Excel-like table (Treeview)
        table_frame = ttk.Frame(parent)
        table_frame.grid(row=2, column=0, columnspan=4, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        columns = ('ID', '任务内容', '重要紧急程度', '完成时间', '状态')
        self.completed_tree = ttk.Treeview(table_frame, columns=columns, show='headings', 
                                          height=20, selectmode='browse')
        
        # Configure columns
        col_widths = {'ID': 50, '任务内容': 400, '重要紧急程度': 120, '完成时间': 150, '状态': 80}
        for col in columns:
            self.completed_tree.heading(col, text=col, anchor=tk.CENTER)
            self.completed_tree.column(col, width=col_widths.get(col, 100), anchor=tk.CENTER)
        
        self.completed_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.completed_tree.yview)
        v_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.completed_tree.configure(yscrollcommand=v_scrollbar.set)
        
        h_scrollbar = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.completed_tree.xview)
        h_scrollbar.grid(row=1, column=0, sticky=(tk.W, tk.E))
        self.completed_tree.configure(xscrollcommand=h_scrollbar.set)
        
        # Configure grid weights
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(2, weight=1)
    
    def add_task(self):
        """Add a new task"""
        task_text = self.task_entry.get().strip()
        if not task_text:
            messagebox.showwarning("警告", "请输入任务内容！")
            return
        
        priority = self.priority_var.get()
        
        task = {
            "id": len(self.data["tasks"]) + len(self.data["completed"]) + 1,
            "text": task_text,
            "priority": priority,
            "created": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "status": "待完成"
        }
        
        self.data["tasks"].append(task)
        self.save_data()
        self.task_entry.delete(0, tk.END)
        self.refresh_display()
        messagebox.showinfo("成功", "任务已添加！")
    
    def complete_task(self):
        """Mark a task as completed"""
        selection = self.pending_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "请选择要完成的任务！")
            return
        
        item = self.pending_tree.item(selection[0])
        task_id = int(item['values'][0])
        
        # Find and move task to completed
        for i, task in enumerate(self.data["tasks"]):
            if task["id"] == task_id:
                task["completed"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                task["status"] = "已完成"
                self.data["completed"].append(task)
                self.data["tasks"].pop(i)
                break
        
        self.save_data()
        self.refresh_display()
        messagebox.showinfo("成功", "任务已标记为完成！")
    
    def delete_task(self):
        """Delete a pending task"""
        selection = self.pending_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "请选择要删除的任务！")
            return
        
        if messagebox.askyesno("确认", "确定要删除这个任务吗？"):
            item = self.pending_tree.item(selection[0])
            task_id = int(item['values'][0])
            
            self.data["tasks"] = [t for t in self.data["tasks"] if t["id"] != task_id]
            self.save_data()
            self.refresh_display()
            messagebox.showinfo("成功", "任务已删除！")
    
    def delete_completed(self):
        """Delete a completed task record"""
        selection = self.completed_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "请选择要删除的记录！")
            return
        
        if messagebox.askyesno("确认", "确定要删除这个记录吗？"):
            item = self.completed_tree.item(selection[0])
            task_id = int(item['values'][0])
            
            self.data["completed"] = [t for t in self.data["completed"] if t["id"] != task_id]
            self.save_data()
            self.refresh_completed()
            messagebox.showinfo("成功", "记录已删除！")
    
    def export_to_excel(self, completed=False):
        """Export tasks to Excel file"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            messagebox.showerror("错误", "请先安装 openpyxl 库！\n使用命令: pip install openpyxl")
            return
        
        # Ask for save location
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="保存Excel文件"
        )
        
        if not file_path:
            return
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "任务列表"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        headers = ['ID', '任务内容', '重要紧急程度', '创建时间', '完成时间', '状态']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Write data
        if completed:
            tasks = self.data["completed"]
        else:
            tasks = self.data["tasks"] + self.data["completed"]
        
        for row, task in enumerate(tasks, 2):
            ws.cell(row=row, column=1, value=task["id"])
            ws.cell(row=row, column=2, value=task["text"])
            ws.cell(row=row, column=3, value=task.get("priority", ""))
            ws.cell(row=row, column=4, value=task["created"])
            ws.cell(row=row, column=5, value=task.get("completed", ""))
            ws.cell(row=row, column=6, value=task.get("status", ""))
            
            # Apply border to all cells in row
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save workbook
        try:
            wb.save(file_path)
            messagebox.showinfo("成功", f"Excel文件已保存至:\n{file_path}")
        except Exception as e:
            messagebox.showerror("错误", f"保存文件时出错:\n{str(e)}")
    
    def refresh_display(self):
        """Refresh the display"""
        self.refresh_pending()
        self.refresh_completed()
    
    def refresh_pending(self):
        """Refresh pending tasks display"""
        self.pending_tree.delete(*self.pending_tree.get_children())
        
        for task in self.data["tasks"]:
            item = self.pending_tree.insert('', tk.END, values=(
                task["id"],
                task["text"],
                task.get("priority", ""),
                task["created"],
                task.get("status", "待完成")
            ))
            
            # Color code based on priority
            priority = task.get("priority", "")
            if priority in self.priority_colors:
                self.pending_tree.set(item, '重要紧急程度', priority)
                # Note: Treeview doesn't easily support cell-level background colors
                # We'll use tags for row-level coloring
                self.pending_tree.item(item, tags=(priority,))
        
        # Configure tag colors
        for priority, color in self.priority_colors.items():
            self.pending_tree.tag_configure(priority, background=color + "40")  # Add transparency
    
    def refresh_completed(self):
        """Refresh completed tasks display"""
        self.completed_tree.delete(*self.completed_tree.get_children())
        
        # Update date filter options
        dates = set()
        for task in self.data["completed"]:
            if "completed" in task:
                date = task["completed"].split()[0]
                dates.add(date)
        
        dates = sorted(list(dates), reverse=True)
        self.date_filter['values'] = ['全部'] + dates
        if not self.date_filter.get():
            self.date_filter.set('全部')
        
        # Display tasks
        self.filter_completed()
    
    def filter_completed(self):
        """Filter completed tasks by date"""
        self.completed_tree.delete(*self.completed_tree.get_children())
        
        selected_date = self.date_filter.get()
        
        for task in self.data["completed"]:
            if "completed" in task:
                task_date = task["completed"].split()[0]
                
                if selected_date == '全部' or task_date == selected_date:
                    item = self.completed_tree.insert('', tk.END, values=(
                        task["id"],
                        task["text"],
                        task.get("priority", ""),
                        task["completed"],
                        task.get("status", "已完成")
                    ))
                    
                    # Color code based on priority
                    priority = task.get("priority", "")
                    if priority in self.priority_colors:
                        self.completed_tree.item(item, tags=(priority,))
        
        # Configure tag colors
        for priority, color in self.priority_colors.items():
            self.completed_tree.tag_configure(priority, background=color + "40")


def main():
    root = tk.Tk()
    app = TodoApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
