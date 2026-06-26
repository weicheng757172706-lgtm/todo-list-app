#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Todo List Application v5.3
Features: Excel-like table, priority matrix, category, person field, persistent data
Layout: pack() based for reliable visibility
"""

VERSION = "6.8"

import json
import os
import sys
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


class TodoApp:

    @staticmethod
    def format_person(person_text):
        """Split multi-person text by common separators, join with ' / ' for display."""
        if not person_text or not person_text.strip():
            return ""
        # Normalize: replace various separators with commas
        for sep in ['，', ';', '；', '、', '|', '/']:
            person_text = person_text.replace(sep, ',')
        # Split, strip, filter empty, join
        names = [n.strip() for n in person_text.split(',') if n.strip()]
        return ' / '.join(names)
    
    def __init__(self, root):
        self.root = root
        self.root.title("ToDoList")
        self.root.geometry("900x650")
        self.root.minsize(700, 500)
        
        # Data file path - save in same directory as exe
        if getattr(sys, 'frozen', False):
            base_path = os.path.dirname(sys.executable)
        else:
            base_path = os.path.dirname(os.path.abspath(__file__))
        
        self.data_file = os.path.join(base_path, "todo_data.json")
        
        # Priority options (Eisenhower Matrix)
        self.priority_options = ["重要紧急", "重要不紧急", "不重要紧急", "不重要不紧急"]
        self.priority_colors = {
            "重要紧急": "#FFCCCC",
            "重要不紧急": "#E5F9F6",
            "不重要紧急": "#FFF9E5",
            "不重要不紧急": "#E5F9E5"
        }

        # Task category options
        self.category_options = ["每日成长任务", "工作任务"]
        
        # Font configurations
        self.font_title = ("Microsoft YaHei", 20, "bold")
        self.font_label = ("Microsoft YaHei", 13)
        self.font_entry = ("Microsoft YaHei", 13)
        self.font_button = ("Microsoft YaHei", 12, "bold")
        self.font_table_header = ("Microsoft YaHei", 12, "bold")
        self.font_table_content = ("Microsoft YaHei", 12)
        
        # Load data
        self.data = self.load_data()
        
        # Setup GUI
        self.setup_gui()
        
        # Refresh display
        self.refresh_display()
    
    def load_data(self):
        """Load data from JSON file"""
        if os.path.exists(self.data_file):
            try:
                with open(self.data_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    if "tasks" not in data:
                        data["tasks"] = []
                    if "completed" not in data:
                        data["completed"] = []
                    return data
            except Exception as e:
                print(f"Error loading data: {e}")
                messagebox.showerror("错误", f"加载数据失败:\n{str(e)}\n将创建新的数据文件。")
        return {"tasks": [], "completed": []}
    
    def save_data(self):
        """Save data to JSON file"""
        try:
            with open(self.data_file, 'w', encoding='utf-8') as f:
                json.dump(self.data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Error saving data: {e}")
            messagebox.showerror("错误", f"保存数据失败:\n{str(e)}")
    
    # ──────────────────────────────────────────────
    #  INPUT AREA — using pack() nested frames
    #  This reliably shows all widgets regardless
    #  of window size.
    # ──────────────────────────────────────────────
    def _build_input_area(self):
        """Build the add-task area using reliable pack() layout."""
        input_frame = tk.LabelFrame(self.root, text="添加新任务",
                                    font=self.font_label,
                                    padx=12, pady=10)
        input_frame.pack(fill=tk.X, padx=15, pady=(10, 5))

        # ── Row 0: task content ──
        row0 = tk.Frame(input_frame)
        row0.pack(fill=tk.X, pady=(0, 8))

        tk.Label(row0, text="任务内容:", font=self.font_label).pack(side=tk.LEFT, padx=(0, 8))

        self.task_entry = tk.Entry(row0, font=self.font_entry)
        self.task_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 12))
        self.task_entry.bind('<Return>', lambda e: self.add_task())

        add_btn = tk.Button(row0, text="✚ 添加任务", command=self.add_task,
                            bg="#4F81BD", fg="white", font=self.font_button,
                            cursor="hand2", padx=20, pady=4)
        add_btn.pack(side=tk.RIGHT)

        # ── Row 1: priority / category / person ──
        row1 = tk.Frame(input_frame)
        row1.pack(fill=tk.X)

        # -- Priority group --
        g1 = tk.Frame(row1)
        g1.pack(side=tk.LEFT, padx=(0, 18))
        tk.Label(g1, text="任务紧急程度:", font=self.font_label).pack(side=tk.LEFT, padx=(0, 6))
        self.priority_var = tk.StringVar(value=self.priority_options[0])
        self.priority_combo = ttk.Combobox(g1, textvariable=self.priority_var,
                                           values=self.priority_options,
                                           state='readonly', width=16, font=self.font_entry)
        self.priority_combo.pack(side=tk.LEFT)

        # -- Category group --
        g2 = tk.Frame(row1)
        g2.pack(side=tk.LEFT, padx=(0, 18))
        tk.Label(g2, text="任务类别:", font=self.font_label).pack(side=tk.LEFT, padx=(0, 6))
        self.category_var = tk.StringVar(value=self.category_options[0])
        self.category_combo = ttk.Combobox(g2, textvariable=self.category_var,
                                           values=self.category_options,
                                           state='readonly', width=14, font=self.font_entry)
        self.category_combo.pack(side=tk.LEFT)

        # -- Person group --
        g3 = tk.Frame(row1)
        g3.pack(side=tk.LEFT)
        tk.Label(g3, text="责任人:", font=self.font_label).pack(side=tk.LEFT, padx=(0, 6))
        self.person_entry = tk.Entry(g3, width=10, font=self.font_entry)
        self.person_entry.pack(side=tk.LEFT)

        return input_frame

    # ──────────────────────────────────────────────
    #  MAIN GUI
    # ──────────────────────────────────────────────
    def setup_gui(self):
        """Setup the GUI components"""
        # Title bar
        title_frame = tk.Frame(self.root, bg="#4F81BD", height=60)
        title_frame.pack(fill=tk.X)
        title_frame.pack_propagate(False)

        title_label = tk.Label(title_frame, text="ToDoList",
                               font=self.font_title,
                               bg="#4F81BD", fg="white")
        title_label.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

        # Input area (pack layout — reliable)
        self._build_input_area()

        # Notebook (tabs)
        notebook = ttk.Notebook(self.root)
        notebook.pack(fill=tk.BOTH, expand=True, padx=15, pady=(5, 10))

        pending_frame = tk.Frame(notebook)
        notebook.add(pending_frame, text="📋 待完成任务")
        self.setup_pending_tab(pending_frame)

        completed_frame = tk.Frame(notebook)
        notebook.add(completed_frame, text="✅ 已完成任务")
        self.setup_completed_tab(completed_frame)

        stats_frame = tk.Frame(notebook)
        notebook.add(stats_frame, text="📊 任务统计")
        self.setup_stats_tab(stats_frame)

        # Status bar
        self.status_bar = tk.Label(self.root, text="就绪", relief=tk.SUNKEN,
                                   anchor=tk.W, font=self.font_label, bg="#F0F0F0")
        self.status_bar.pack(fill=tk.X, padx=15, pady=(0, 10))
    
    def setup_pending_tab(self, parent):
        """Setup pending tasks tab"""
        toolbar = tk.Frame(parent)
        toolbar.pack(fill=tk.X, pady=(10, 10))
        
        tk.Button(toolbar, text="✓ 标记完成", command=self.complete_task,
                  bg="#5CB85C", fg="white", font=self.font_button,
                  cursor="hand2", padx=16, pady=4).pack(side=tk.LEFT, padx=(0, 8))
        
        tk.Button(toolbar, text="🗑 删除任务", command=self.delete_task,
                  bg="#D9534F", fg="white", font=self.font_button,
                  cursor="hand2", padx=16, pady=4).pack(side=tk.LEFT, padx=(0, 8))

        tk.Button(toolbar, text="✏ 编辑", command=self.edit_task,
                  bg="#F0AD4E", fg="white", font=self.font_button,
                  cursor="hand2", padx=16, pady=4).pack(side=tk.LEFT, padx=(0, 8))
        
        # Table
        table_frame = tk.Frame(parent)
        table_frame.pack(fill=tk.BOTH, expand=True)
        
        columns = ('ID', '任务内容', '任务类别', '任务紧急程度', '责任人', '创建时间')
        self.pending_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=20)
        
        style = ttk.Style()
        style.configure("Treeview", font=self.font_table_content, rowheight=45)
        style.configure("Treeview.Heading", font=self.font_table_header)
        
        col_widths = {'ID': 45, '任务内容': 310, '任务类别': 120, '任务紧急程度': 130, '责任人': 140, '创建时间': 170}
        for col in columns:
            self.pending_tree.heading(col, text=col, anchor=tk.CENTER)
            stretch = (col == '任务内容')
            anchor = tk.W if col == '责任人' else tk.CENTER
            self.pending_tree.column(col, width=col_widths.get(col, 100),
                                    minwidth=40, stretch=stretch, anchor=anchor)
        
        vsb = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.pending_tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.pending_tree.xview)
        self.pending_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.pending_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
    
    def setup_completed_tab(self, parent):
        """Setup completed tasks tab"""
        # Filter
        filter_frame = tk.Frame(parent)
        filter_frame.pack(fill=tk.X, pady=(10, 10))

        tk.Label(filter_frame, text="筛选日期:", font=self.font_label).pack(side=tk.LEFT, padx=(0, 8))

        self.date_filter = ttk.Combobox(filter_frame, width=22, state='readonly', font=self.font_entry)
        self.date_filter.pack(side=tk.LEFT, padx=(0, 15))
        self.date_filter.bind('<<ComboboxSelected>>', lambda e: self.filter_completed())

        # Toolbar
        toolbar = tk.Frame(parent)
        toolbar.pack(fill=tk.X, pady=(0, 10))

        tk.Button(toolbar, text="🗑 删除所选", command=self.delete_completed,
                  bg="#E67E22", fg="white", font=self.font_button,
                  cursor="hand2", padx=16, pady=4).pack(side=tk.LEFT, padx=(0, 8))

        tk.Button(toolbar, text="🗑 删除全部", command=self.delete_all_completed,
                  bg="#C0392B", fg="white", font=self.font_button,
                  cursor="hand2", padx=16, pady=4).pack(side=tk.LEFT, padx=(0, 16))

        tk.Button(toolbar, text="📤 导出 Excel", command=self.export_completed,
                  bg="#4F81BD", fg="white", font=self.font_button,
                  cursor="hand2", padx=16, pady=4).pack(side=tk.LEFT, padx=(0, 8))

        tk.Label(toolbar, text="已完成任务自动保存，关闭后不丢失",
                 font=self.font_label, fg="#888888").pack(side=tk.LEFT, padx=(20, 0))

        # Table
        table_frame = tk.Frame(parent)
        table_frame.pack(fill=tk.BOTH, expand=True)

        columns = ('ID', '任务内容', '任务类别', '任务紧急程度', '责任人', '完成时间')
        self.completed_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=20)
        
        col_widths = {'ID': 45, '任务内容': 310, '任务类别': 120, '任务紧急程度': 130, '责任人': 140, '完成时间': 170}
        for col in columns:
            self.completed_tree.heading(col, text=col, anchor=tk.CENTER)
            stretch = (col == '任务内容')
            anchor = tk.W if col == '责任人' else tk.CENTER
            self.completed_tree.column(col, width=col_widths.get(col, 100),
                                      minwidth=40, stretch=stretch, anchor=anchor)

        vsb = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.completed_tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.completed_tree.xview)
        self.completed_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.completed_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
    
    def setup_stats_tab(self, parent):
        """Setup statistics tab with overview and breakdown tables"""
        # ── 0. Week filter ──
        filter_frame = tk.Frame(parent)
        filter_frame.pack(fill=tk.X, padx=10, pady=(10, 5))
        
        tk.Label(filter_frame, text="选择周:", font=self.font_label).pack(side=tk.LEFT, padx=(0, 8))
        
        self.week_filter_var = tk.StringVar(value="全部")
        self.week_filter_combo = ttk.Combobox(filter_frame, textvariable=self.week_filter_var,
                                               width=22, state='readonly', font=self.font_entry)
        self.week_filter_combo.pack(side=tk.LEFT, padx=(0, 15))
        self.week_filter_combo.bind('<<ComboboxSelected>>',
                                    lambda e: self.refresh_stats())
        
        # ── 1. Overview cards ──
        overview_frame = tk.LabelFrame(parent, text="📊 总览",
                                       font=self.font_label, padx=15, pady=12)
        overview_frame.pack(fill=tk.X, padx=10, pady=(0, 8))
        
        card_frame = tk.Frame(overview_frame)
        card_frame.pack(fill=tk.X)
        
        self.stat_pending = tk.StringVar(value="0")
        self.stat_completed = tk.StringVar(value="0")
        self.stat_rate = tk.StringVar(value="0%")
        
        card_data = [
            ("待完成任务", self.stat_pending, "#F0AD4E"),
            ("已完成任务", self.stat_completed, "#5CB85C"),
            ("完成率", self.stat_rate, "#4F81BD"),
        ]
        for i, (label, var, bg) in enumerate(card_data):
            card = tk.Frame(card_frame, bg=bg, padx=20, pady=8)
            card.pack(side=tk.LEFT, fill=tk.X, expand=True,
                      padx=(0, 10) if i < len(card_data) - 1 else 0)
            tk.Label(card, text=label, font=self.font_label,
                     bg=bg, fg="white").pack()
            tk.Label(card, textvariable=var,
                     font=("Microsoft YaHei", 28, "bold"),
                     bg=bg, fg="white").pack()
        
        # ── 2. By Priority + By Category side by side ──
        mid_frame = tk.Frame(parent)
        mid_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
        
        # -- Priority table --
        pri_frame = tk.LabelFrame(mid_frame, text="按紧急程度",
                                  font=self.font_label, padx=8, pady=8)
        pri_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        
        p_cols = ('紧急程度', '待完成', '已完成', '合计')
        self.stat_priority_tree = ttk.Treeview(pri_frame, columns=p_cols,
                                               show='headings', height=5)
        for c in p_cols:
            self.stat_priority_tree.heading(c, text=c, anchor=tk.CENTER)
            self.stat_priority_tree.column(c, width=95, minwidth=70, anchor=tk.CENTER)
        self.stat_priority_tree.pack(fill=tk.BOTH, expand=True)
        
        # -- Category table --
        cat_frame = tk.LabelFrame(mid_frame, text="按任务类别",
                                  font=self.font_label, padx=8, pady=8)
        cat_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        
        c_cols = ('任务类别', '待完成', '已完成', '合计')
        self.stat_category_tree = ttk.Treeview(cat_frame, columns=c_cols,
                                               show='headings', height=5)
        for c in c_cols:
            self.stat_category_tree.heading(c, text=c, anchor=tk.CENTER)
            self.stat_category_tree.column(c, width=105, minwidth=80, anchor=tk.CENTER)
        self.stat_category_tree.pack(fill=tk.BOTH, expand=True)
        
        # ── 3. By Date ──
        date_frame = tk.LabelFrame(parent, text="按日期统计（已完成）",
                                   font=self.font_label, padx=8, pady=8)
        date_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        
        d_cols = ('日期', '完成数')
        self.stat_date_tree = ttk.Treeview(date_frame, columns=d_cols,
                                           show='headings')
        for c in d_cols:
            self.stat_date_tree.heading(c, text=c, anchor=tk.CENTER)
            self.stat_date_tree.column(c, width=160, minwidth=100, anchor=tk.CENTER)
        self.stat_date_tree.pack(fill=tk.BOTH, expand=True)
    
    def add_task(self):
        """Add a new task"""
        task_text = self.task_entry.get().strip()
        if not task_text:
            messagebox.showwarning("警告", "请输入任务内容！")
            return
        
        priority = self.priority_var.get()
        category = self.category_var.get()
        person = self.person_entry.get().strip()
        
        task = {
            "id": len(self.data["tasks"]) + len(self.data["completed"]) + 1,
            "text": task_text,
            "priority": priority,
            "category": category,
            "person": person,
            "created": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        
        self.data["tasks"].append(task)
        self.save_data()
        self.task_entry.delete(0, tk.END)
        self.person_entry.delete(0, tk.END)
        self.refresh_display()
        self.status_bar.config(text=f"✓ 任务已添加: {task_text}")
    
    def complete_task(self):
        """Mark a task as completed"""
        selection = self.pending_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "请选择要完成的任务！")
            return
        
        task_id = int(selection[0])
        
        for i, task in enumerate(self.data["tasks"]):
            if task["id"] == task_id:
                task["completed"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                self.data["completed"].append(task)
                self.data["tasks"].pop(i)
                break
        
        self.save_data()
        self.refresh_display()
        self.status_bar.config(text="✓ 任务已标记为完成")
    
    def delete_task(self):
        """Delete a pending task"""
        selection = self.pending_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "请选择要删除的任务！")
            return
        
        task_id = int(selection[0])
        self.data["tasks"] = [t for t in self.data["tasks"] if t["id"] != task_id]
        self.save_data()
        self.refresh_display()
        self.status_bar.config(text="✓ 任务已删除")
    
    def edit_task(self):
        """Edit selected pending task's content, category, priority, or person"""
        selection = self.pending_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "请选择要编辑的任务！")
            return
        
        task_id = int(selection[0])
        task = None
        for t in self.data["tasks"]:
            if t["id"] == task_id:
                task = t
                break
        
        if not task:
            messagebox.showerror("错误", "未找到该任务！")
            return
        
        # ── Build edit dialog ──
        dialog = tk.Toplevel(self.root)
        dialog.title(f"编辑任务 #{task_id}")
        dialog.geometry("520x300")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        
        main_frame = tk.Frame(dialog, padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # --- 任务内容 ---
        tk.Label(main_frame, text="任务内容:", font=self.font_label).pack(anchor=tk.W, pady=(0, 4))
        text_var = tk.StringVar(value=task.get("text", ""))
        text_entry = tk.Entry(main_frame, font=self.font_entry, textvariable=text_var)
        text_entry.pack(fill=tk.X, pady=(0, 14))
        text_entry.focus_set()
        
        # --- 任务类别 ---
        row_frame = tk.Frame(main_frame)
        row_frame.pack(fill=tk.X, pady=(0, 14))
        
        g1 = tk.Frame(row_frame)
        g1.pack(side=tk.LEFT, padx=(0, 20))
        tk.Label(g1, text="任务类别:", font=self.font_label).pack(anchor=tk.W, pady=(0, 4))
        category_var = tk.StringVar(value=task.get("category", self.category_options[0]))
        category_combo = ttk.Combobox(g1, textvariable=category_var,
                                      values=self.category_options,
                                      state='readonly', width=16, font=self.font_entry)
        category_combo.pack()
        
        g2 = tk.Frame(row_frame)
        g2.pack(side=tk.LEFT)
        tk.Label(g2, text="任务紧急程度:", font=self.font_label).pack(anchor=tk.W, pady=(0, 4))
        priority_var = tk.StringVar(value=task.get("priority", self.priority_options[0]))
        priority_combo = ttk.Combobox(g2, textvariable=priority_var,
                                      values=self.priority_options,
                                      state='readonly', width=16, font=self.font_entry)
        priority_combo.pack()
        
        # --- 责任人 ---
        tk.Label(main_frame, text="责任人:", font=self.font_label).pack(anchor=tk.W, pady=(0, 4))
        person_var = tk.StringVar(value=task.get("person", ""))
        person_entry = tk.Entry(main_frame, font=self.font_entry, textvariable=person_var)
        person_entry.pack(fill=tk.X, pady=(0, 18))
        
        # --- Buttons ---
        btn_frame = tk.Frame(main_frame)
        btn_frame.pack(fill=tk.X)
        
        def save_edit():
            new_text = text_var.get().strip()
            if not new_text:
                messagebox.showwarning("警告", "任务内容不能为空！")
                return
            task["text"] = new_text
            task["category"] = category_var.get()
            task["priority"] = priority_var.get()
            task["person"] = person_var.get().strip()
            self.save_data()
            self.refresh_display()
            self.status_bar.config(text=f"✓ 任务已更新: {new_text}")
            dialog.destroy()
        
        tk.Button(btn_frame, text="✚ 保存修改", command=save_edit,
                  bg="#4F81BD", fg="white", font=self.font_button,
                  cursor="hand2", padx=20, pady=4).pack(side=tk.RIGHT, padx=(10, 0))
        
        tk.Button(btn_frame, text="取消", command=dialog.destroy,
                  font=self.font_button,
                  cursor="hand2", padx=20, pady=4).pack(side=tk.RIGHT)
    
    def delete_completed(self):
        """Delete a completed task record"""
        selection = self.completed_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "请选择要删除的记录！")
            return
        
        task_id = int(selection[0])
        self.data["completed"] = [t for t in self.data["completed"] if t["id"] != task_id]
        self.save_data()
        self.refresh_completed()
        self.status_bar.config(text="✓ 记录已删除")
    
    def delete_all_completed(self):
        """Delete all completed task records"""
        count = len(self.data["completed"])
        if count == 0:
            self.status_bar.config(text="已完成列表为空，无需删除")
            return
        self.data["completed"] = []
        self.save_data()
        self.refresh_completed()
        self.refresh_display()
    
    def export_completed(self):
        """Export all completed tasks to Excel file"""
        if not self.data["completed"]:
            self.status_bar.config(text="已完成列表为空，无可导出的记录")
            return
        
        now = datetime.now()
        iso_year, iso_week, _ = now.isocalendar()
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
            title="导出已完成任务",
            initialfile=f"{iso_year}年第{iso_week}周 已完成任务列表.xlsx"
        )
        if not file_path:
            return  # user cancelled
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "已完成任务"
        
        # ── Header style ──
        header_font = Font(name="Microsoft YaHei", bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        
        # ── Write headers ──
        headers = ["序号", "任务内容", "任务类别", "任务紧急程度", "责任人", "完成时间", "创建时间"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        
        # ── Write data ──
        data_font = Font(name="Microsoft YaHei", size=11)
        data_align = Alignment(horizontal="center", vertical="center")
        content_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        for row_idx, task in enumerate(self.data["completed"], 2):
            values = [
                row_idx - 1,
                task.get("text", ""),
                task.get("category", ""),
                task.get("priority", ""),
                self.format_person(task.get("person", "")),
                task.get("completed", ""),
                task.get("created", "")
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.border = thin_border
                if col_idx == 2:  # 任务内容 - left align
                    cell.alignment = content_align
                else:
                    cell.alignment = data_align
        
        # ── Auto-fit column widths ──
        col_widths_map = {1: 6, 2: 45, 3: 14, 4: 14, 5: 16, 6: 18, 7: 18}
        for col_idx, width in col_widths_map.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
        
        # Freeze top row
        ws.freeze_panes = "A2"
        
        try:
            wb.save(file_path)
            self.status_bar.config(text=f"✓ 导出成功: {os.path.basename(file_path)} ({len(self.data['completed'])} 条记录)")
        except Exception as e:
            messagebox.showerror("导出失败", f"保存文件失败:\n{str(e)}")
    
    def refresh_display(self):
        """Refresh the display"""
        self.refresh_pending()
        self.refresh_completed()
        self.refresh_stats()
        self.status_bar.config(text=f"就绪 | 待完成: {len(self.data['tasks'])} | 已完成: {len(self.data['completed'])}")
    
    def refresh_pending(self):
        """Refresh pending tasks display"""
        self.pending_tree.delete(*self.pending_tree.get_children())
        
        for display_id, task in enumerate(self.data["tasks"], 1):
            item = self.pending_tree.insert('', tk.END, iid=str(task["id"]), values=(
                display_id,
                task["text"],
                task.get("category", ""),
                task.get("priority", ""),
                self.format_person(task.get("person", "")),
                task["created"]
            ))
            
            priority = task.get("priority", "")
            if priority in self.priority_colors:
                self.pending_tree.item(item, tags=(priority,))
        
        for priority, color in self.priority_colors.items():
            self.pending_tree.tag_configure(priority, background=color)
    
    def refresh_completed(self):
        """Refresh completed tasks display"""
        self.completed_tree.delete(*self.completed_tree.get_children())
        
        dates = set()
        for task in self.data["completed"]:
            if "completed" in task:
                date = task["completed"].split()[0]
                dates.add(date)
        
        dates = sorted(list(dates), reverse=True)
        self.date_filter['values'] = ['全部'] + dates
        if not self.date_filter.get():
            self.date_filter.set('全部')
        
        self.filter_completed()
    
    def filter_completed(self):
        """Filter completed tasks by date"""
        self.completed_tree.delete(*self.completed_tree.get_children())
        
        selected_date = self.date_filter.get()
        display_id = 0
        
        for task in self.data["completed"]:
            if "completed" in task:
                task_date = task["completed"].split()[0]
                
                if selected_date == '全部' or task_date == selected_date:
                    display_id += 1
                    item = self.completed_tree.insert('', tk.END, iid=str(task["id"]), values=(
                        display_id,
                        task["text"],
                        task.get("category", ""),
                        task.get("priority", ""),
                        self.format_person(task.get("person", "")),
                        task["completed"]
                    ))
                    
                    priority = task.get("priority", "")
                    if priority in self.priority_colors:
                        self.completed_tree.item(item, tags=(priority,))
        
        for priority, color in self.priority_colors.items():
            self.completed_tree.tag_configure(priority, background=color)
    
    # ──────────────────────────────────────────────
    #  STATISTICS — Table-based, week-aware
    # ──────────────────────────────────────────────
    def _date_to_week(self, date_str):
        """Parse a date string (YYYY-MM-DD ...) and return (iso_year, iso_week)."""
        try:
            d = datetime.strptime(date_str.split()[0], "%Y-%m-%d")
            iso = d.isocalendar()
            return (iso[0], iso[1])
        except (ValueError, IndexError):
            return None

    def _build_week_options(self):
        """Build week dropdown options from all task dates."""
        weeks_set = set()
        for t in self.data["tasks"]:
            w = self._date_to_week(t.get("created", ""))
            if w:
                weeks_set.add(w)
        for t in self.data["completed"]:
            w = self._date_to_week(t.get("completed", ""))
            if w:
                weeks_set.add(w)
        self._all_weeks = sorted(weeks_set, reverse=True)
        options = ["全部"] + [f"{y}年第{w}周" for y, w in self._all_weeks]
        self.week_filter_combo['values'] = options
        if not self.week_filter_var.get() or self.week_filter_var.get() not in options:
            self.week_filter_var.set(options[0] if options else "全部")

    def _filter_by_week(self, tasks, date_field):
        """Filter a list of tasks by the currently selected week.
        Returns filtered list, or original list if selected week is '全部'.
        """
        selected = self.week_filter_var.get()
        if selected == "全部":
            return tasks
        import re
        m = re.match(r"(\d+)年.*?(\d+)周", selected)
        if not m:
            return tasks
        year_filter, week_filter = int(m.group(1)), int(m.group(2))
        result = []
        for t in tasks:
            w = self._date_to_week(t.get(date_field, ""))
            if w and w[0] == year_filter and w[1] == week_filter:
                result.append(t)
        return result

    def refresh_stats(self):
        """Refresh statistics tab data, filtered by selected week"""
        self._build_week_options()
        
        # Filter by week
        filtered_tasks = self._filter_by_week(self.data["tasks"], "created")
        filtered_completed = self._filter_by_week(self.data["completed"], "completed")
        
        total_pending = len(filtered_tasks)
        total_completed = len(filtered_completed)
        total_all = total_pending + total_completed
        
        self.stat_pending.set(str(total_pending))
        self.stat_completed.set(str(total_completed))
        self.stat_rate.set(f"{total_completed / total_all * 100:.1f}%" if total_all > 0 else "0%")
        
        # ── By priority ──
        self.stat_priority_tree.delete(*self.stat_priority_tree.get_children())
        for pri in self.priority_options:
            p = sum(1 for t in filtered_tasks if t.get("priority") == pri)
            c = sum(1 for t in filtered_completed if t.get("priority") == pri)
            total = p + c
            if p > 0 or c > 0:
                self.stat_priority_tree.insert('', tk.END, values=(pri, p, c, total))
        
        # ── By category ──
        self.stat_category_tree.delete(*self.stat_category_tree.get_children())
        for cat in self.category_options:
            p = sum(1 for t in filtered_tasks if t.get("category") == cat)
            c = sum(1 for t in filtered_completed if t.get("category") == cat)
            total = p + c
            if p > 0 or c > 0:
                self.stat_category_tree.insert('', tk.END, values=(cat, p, c, total))
        
        # ── By date ──
        self.stat_date_tree.delete(*self.stat_date_tree.get_children())
        date_counts = {}
        for t in filtered_completed:
            d = t.get("completed", "").split()[0]
            date_counts[d] = date_counts.get(d, 0) + 1
        for date in sorted(date_counts.keys(), reverse=True):
            self.stat_date_tree.insert('', tk.END, values=(date, date_counts[date]))


def main():
    root = tk.Tk()
    app = TodoApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
