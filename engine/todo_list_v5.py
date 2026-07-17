#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Todo List Application v5.3
Features: Excel-like table, priority matrix, category, person field, persistent data
Layout: pack() based for reliable visibility
"""

VERSION = "V1.9.7"  # 版本号按 V.A.B.C 新规（逢10进1）。V1.9.7 彻底移除独立「成就解锁」系统（窗口+判定+数据），仅保留成长系统(等级/经验/称号)；双剑/清单同步升版

import json
import os
import sys
import copy
import re
import socket
import shutil
from datetime import datetime, date, timedelta
import math
import random
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# ══════════════════════════════════════════════════════════════
# 主题系统（多套主题架构：单引擎 + theme.json 配置）
# 默认值为稳重风（冒险双剑）。打包时 --add-data 注入各套 theme.json
# 即可生成不同皮肤、独立运行的 exe。emoji 本期统一灰版，预留 theme.emoji。
# ══════════════════════════════════════════════════════════════
DEFAULT_THEME = {
    "name": "冒险游戏",
    "title": "冒险者の游戏",
    "icon": "TodoList_icon.ico",
    "port": 54321,
    "colors": {
        # 观感骨架色（换肤主要体现处，源码引用 THEME['colors'][key]）
        "title_bar": "#2C3E50",
        "status_bar": "#ECF0F1",
        "status_fg": "#333333",
        "panel_bg": "#F0F0F0",
        "primary": "#4F81BD",
        "accent": "#F1C40F",
        "dark_alt": "#34495E",
        "dark_alt2": "#1B4F72",
        "exp_card_bg": "#2C3E50",
        "exp_card_fg": "#ECF0F1",
        "exp_gold": "#F1C40F",
        "exp_green": "#2ECC71",
        "exp_sub": "#BDC3C7",
        "clock_border": "#4F81BD",
        "clock_bg": "#F0F0F0",
        "clock_fg": "#2C3E50",
        "tree_selected": "#F0F0F0",
        "tree_subtask": "#F5F5F5",
        "tree_subtask_done": "#E8E8E8",
        "done_today": "#A9A9A9",
        "done_old": "#D3D3D3",
        "progress_track": "#34495E",
        "progress_accent": "#F1C40F",
        # 功能/状态语义色（跨主题保持一致，源码保持 hex 字面量，不抽主题）
        "success": "#5CB85C",
        "danger": "#D9534F",
        "warning": "#F0AD4E",
        "muted": "#6C757D",
        "info": "#3498DB",
        "muted_alt": "#95A5A6",
        "warning_alt": "#E67E22",
        "danger_alt": "#C0392B",
        "warning_alt2": "#F39C12",
        "danger_alt2": "#E74C3C",
        "gold_alt": "#FFD700",
        "success_alt": "#00AA00",
        "title_color": "#2C3E50",
        "hint": "#888888",
        "purple": "#9B59B6",
        "dark_bg": "#2B2B2B",
        "dark_fg": "#9B59B6",
    },
    # ── 界面文案（多套主题差异化核心；emoji 随文字一起主题化）──
    # 默认=冒险双剑文案。清单打勾等套在 theme.json 的 text 段覆盖即可。
    "text": {
        # 列名（含 emoji；同时作为 Treeview 列 key 与显示一体）
        "col_id": "编号",
        "col_id_emoji": "🔢 编号",
        "col_desc": "📝 冒险描述",
        "col_type": "🗺️ 冒险类型",
        "col_priority": "⚡ 紧急等级",
        "col_owner": "⚔️ 冒险者",
        "col_tag": "🏷️ 标签",
        "form_tag": "🏷️ 标签:",
        "dlg_tag": "🏷️ 标签 (逗号分隔):",
        "dlg_tag2": "🏷️ 标签: (逗号分隔)",
        "col_progress": "📊 进度",
        "col_receive": "⏰ 领取时间",
        "col_complete_time": "🏁 通关时间",
        "col_duration": "⏱️ 冒险时长",
        # 顶部 tab
        "tab_pending": "🚀 进行中的冒险",
        "tab_completed": "🏁 已通关的冒险",
        "tab_stats": "📊 冒险战绩",
        # 工具栏按钮
        "btn_mark": "🎯 标记通关",
        "btn_abandon": "🗑 放弃冒险",
        "btn_edit": "✏ 编辑冒险",
        "btn_refresh": "🔄 刷新冒险",
        "btn_clear": "🗑 清空冒险",
        "btn_export": "📤 导出战绩",
        # 表单标签
        "lbl_desc_add": "🎯 冒险描述:",
        "lbl_desc_edit": "📝 冒险描述:",
        "lbl_type": "🗺️ 冒险类型:",
        "lbl_priority": "⚡ 紧急等级:",
        "lbl_owner": "⚔️ 冒险者:",
        # 统计区
        "stat_priority": "⚡ 按紧急等级",
        "stat_category": "🗺️ 按冒险类型",
        "stat_ongoing": "🚀 进行中",
        "stat_done": "🏁 已通关",
        "stat_total": "📊 合计",
        # 任务大厅总览卡片（3 张）
        "overview_pending": "🚀 进行中",
        "overview_done": "🏁 已通关",
        "overview_rate": "📊 通关率",
        # 排序下拉
        "sort_default": "默认顺序",
        "sort_priority": "按紧急程度",
        "sort_created": "按创建时间",
        "sort_desc": "按任务描述",
        "sort_completed": "按通关时间",
        # 导出
        "export_title": "📤 导出战绩",
        "export_sheet": "🏁 已通关冒险",
        "export_file_week": "{week} 已通关冒险列表.xlsx",
        "export_file_iso": "{year}年第{week}周 已通关冒险列表.xlsx",
        # 状态栏 / 校验 / 弹窗
        "status_enter_desc": "🎯 请输入冒险描述...",
        "warn_enter_desc": "🎯 请输入冒险描述！",
        "valid_owner": "冒险者名称不能超过100个字符！",
        "edit_dlg_title": "✏️ 编辑冒险 #",
        # ── 经验/日期/成就等（原硬编码，抽主题以便清单套去冒险味）──
        "exp_card_title": "🏆 本周黄金之路",
        "exp_daily": "📅 本周第{}天，剩余{}天",
        "date_stat_title": "📅 按日期统计（已通关）",
        "date_col": "📅 日期",
        "done_count_col": "🏁 通关数",
        "levelup_title": "🎉 等级提升！",
        "levelup_msg": "等级提升！",
        "levelup_icon": "🎉",
        "levelup_btn": "继续冒险！",
        "complete_msg": "🎉 恭喜您冒险已完成！！！",
        "complete_emoji": "🎉",
        "complete_title": "恭喜您，冒险已完成！",
        "complete_sub": "本次冒险已记入战绩",
        "complete_chip": "+10 经验值",
        "celebrate_accent": "#F1C40F",
        "chip_bg": "#F1C40F",
        "chip_fg": "#2C3E50",
        "wisdom_anim": "🧠 智慧 +{} 🧠",
        # ── 状态栏 / 对话框 / 右键菜单 / 提示（原硬编码，抽主题以便清单套去冒险味）──
        "msg_new_task_gen": "🎯 新冒险已生成！",
        "btn_add_task": "✚ 领取新冒险",
        "status_ready": "准备开始新的冒险！",
        "status_ready_icon": "⚔️ 准备开始新的冒险！",
        "ctx_new_task": "🎯 领取新冒险",
        "ctx_add_subtask": "🎯 接取支线任务",
        "warn_select_sub": "请选择要接取支线的冒险！",
        "err_not_found": "未找到该冒险！",
        "warn_select_sub_manage": "请选择要管理支线任务的冒险！",
        "btn_del_completed": "🗑 删除冒险",
        "ctx_restore": "🔄 恢复冒险",
        "ctx_del_completed": "🗑 删除冒险",
        "status_task_received": "🎯 冒险已领取：{}",
        "err_new_task": "🎯 领取新冒险出错",
        "warn_select_complete": "🎮 请选择要通关的冒险！",
        "status_task_done": "🏁 冒险已通关！",
        "warn_select_abandon": "🗑️ 请选择要放弃的冒险！",
        "confirm_abandon": "🗑️ 确定要放弃这个冒险吗？\n\n冒险: {}",
        "status_abandoned": "🗑️ 冒险已放弃！",
        "warn_select_edit": "🎮 请选择要编辑的冒险！",
        "err_not_found_edit": "🎮 未找到该冒险！",
        "status_task_updated": "✏️ 冒险已更新: {}",
        "warn_select_restore": "🔄 请选择要恢复的冒险！",
        "status_task_restored": "🔄 冒险已恢复：{}（扣回 {} 智慧）",
        "confirm_del_done": "🗑️ 确定要删除该已通关记录吗？\n\n冒险: {}\n\n⚠️ 此操作将扣回 {} 智慧（历史总智慧同步减少），本周等级可能下降",
        "dlg_title_add": "🎯 接取新冒险",
        "dlg_title_receive": "🎯 领取新冒险",
        "btn_save_new": "✅ 领取冒险",
    },
    # ── 成长级别（Lv + 称号 + 经验阈值；阈值同原冒险体系，称号可按套覆盖）──
    "levels": [
        {"name": "菜鸟路人", "min": 0, "max": 60},
        {"name": "见习旅者", "min": 61, "max": 200},
        {"name": "荒野智者", "min": 201, "max": 300},
        {"name": "探险泰斗", "min": 301, "max": 400},
        {"name": "传奇冒险家", "min": 401, "max": 500},
        {"name": "至尊冒险王", "min": 501, "max": 999999},
    ],
    # ── 下拉选项（优先级/类型；按套覆盖，整块替换非深合并）──
    # 注意：经验值/颜色/排序权重/加粗 tag 均按「位置索引」与档位名解耦，避免硬编码档位名。
    "priority_options": ["红色警报", "修炼升级", "临时密令", "佛系摸鱼"],
    "category_options": ["主线任务", "营地修行"],
    # 旧档位名→当前档位名 迁移映射（双剑套默认；清单套在 theme.json 整块覆盖为反向）
    "legacy_priority_map": {
        "重要紧急": "红色警报", "重要不紧急": "修炼升级", "不重要紧急": "临时密令",
        "不重要不紧急": "佛系摸鱼", "紧急Boss战": "红色警报", "主线任务": "修炼升级",
        "突发委托": "临时密令", "摸鱼任务": "佛系摸鱼", "游击战": "佛系摸鱼"
    },
    "legacy_category_map": {
        "工作任务": "主线任务", "每日任务": "营地修行", "重要冒险": "主线任务",
        "每日修行": "营地修行", "搬砖任务": "主线任务"
    },
}


def _deep_merge(base, override):
    """递归合并 override 到 base（depth-first），返回 base。"""
    for k, v in override.items():
        if isinstance(v, dict) and isinstance(base.get(k), dict):
            _deep_merge(base[k], v)
        else:
            base[k] = v
    return base


def load_theme():
    """加载主题：优先 _MEIPASS（打包后），其次脚本同目录的 theme.json。
    缺失字段回退 DEFAULT_THEME，保证无 theme.json 时行为不变（稳重风）。"""
    theme = copy.deepcopy(DEFAULT_THEME)
    candidates = []
    if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
        candidates.append(os.path.join(sys._MEIPASS, 'theme.json'))
    candidates.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'theme.json'))
    for p in candidates:
        if os.path.exists(p):
            try:
                with open(p, 'r', encoding='utf-8') as f:
                    override = json.load(f)
                _deep_merge(theme, override)
                # legacy 迁移映射整块覆盖（dict 深合并会叠加，需显式整块替换）
                for _k in ('legacy_priority_map', 'legacy_category_map'):
                    if _k in override:
                        theme[_k] = override[_k]
            except Exception:
                pass
            break
    return theme


THEME = load_theme()

# ──────────────────────────────────────────────
# Single instance lock (prevent multiple instances)
# ──────────────────────────────────────────────
LOCK_PORT = THEME['port']  # 每套主题独立端口，避免多套同时运行互相锁

def is_already_running():
    """Check if another instance is already running using a socket lock."""
    global _lock_socket
    _lock_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        _lock_socket.bind(("127.0.0.1", LOCK_PORT))
        return False
    except socket.error:
        return True

class TodoApp:

    @staticmethod
    def format_person(person_text):
        """Split multi-person text by common separators, join with ' / ' for display."""
        if not person_text or person_text is None:
            return ""
        person_text = str(person_text).strip()
        if not person_text:
            return ""
        # Normalize: replace various separators with commas
        for sep in ['，', ';', '；', '、', '|', '/']:
            person_text = person_text.replace(sep, ',')
        # Split, strip, filter empty, join
        names = [n.strip() for n in person_text.split(',') if n.strip()]
        return ' / '.join(names)
    
    @staticmethod
    def validate_task_text(text):
        """Validate task description.
        Returns: (is_valid, error_message)
        """
        if not text or not text.strip():
            return False, "任务描述不能为空！"
        
        text = text.strip()
        
        # Check length
        if len(text) > 500:
            return False, "任务描述不能超过500个字符！"
        
        # Check for potentially dangerous characters
        dangerous_chars = ['<', '>', '{', '}', '[', ']', '|', '\\', '`']
        for char in dangerous_chars:
            if char in text:
                return False, f"任务描述不能包含字符: {char}"
        
        return True, ""
    
    def validate_person_name(self, person):
        """Validate person name.
        Returns: (is_valid, error_message)
        """
        if not person or not person.strip():
            return True, ""  # Person is optional, default to "韦程"
        
        person = person.strip()
        
        # Check length
        if len(person) > 100:
            return False, self.T['valid_owner']
        
        return True, ""
    
    @staticmethod
    def validate_tags(tags_text):
        """Validate tags text.
        Returns: (is_valid, error_message, tags_list)
        """
        if not tags_text or not tags_text.strip():
            return True, "", []
        
        # Split by comma and validate each tag
        tags = [tag.strip() for tag in tags_text.split(',') if tag.strip()]
        
        # Check number of tags
        if len(tags) > 20:
            return False, "标签数量不能超过20个！", []
        
        # Check each tag
        for tag in tags:
            if len(tag) > 50:
                return False, f"标签 '{tag}' 不能超过50个字符！", []
            if not tag.isprintable():
                return False, f"标签 '{tag}' 包含不可打印字符！", []
        
        return True, "", tags
    
    @staticmethod
    def strikethrough(text):
        """Add strikethrough effect to text using Markdown-style markers."""
        if not text:
            return ""
        return f"~~{text}~~"

    @staticmethod
    def calculate_duration(created_str, completed_str):
        """Calculate time difference between created and completed timestamps.
        Returns formatted string like 'X天X小时X分钟'."""
        try:
            created = datetime.strptime(created_str, "%Y-%m-%d %H:%M:%S")
            completed = datetime.strptime(completed_str, "%Y-%m-%d %H:%M:%S")
            delta = completed - created
            
            total_seconds = int(delta.total_seconds())
            days = total_seconds // 86400
            hours = (total_seconds % 86400) // 3600
            minutes = (total_seconds % 3600) // 60
            
            parts = []
            if days > 0:
                parts.append(f"{days}天")
            if hours > 0:
                parts.append(f"{hours}小时")
            if minutes > 0 or len(parts) == 0:
                parts.append(f"{minutes}分钟")
            
            return "".join(parts)
        except (ValueError, TypeError):
            return "未知"
    
    def __init__(self, root):
        self.root = root
        # 主题化文案 / 等级（多套主题：双剑用默认冒险文案，清单套覆盖）
        self.T = THEME['text']
        self.LEVELS = THEME['levels']
        self.root.title(f"{THEME['title']} {VERSION}")
        # ── 主窗口尺寸：首开与「从最大化还原」共用唯一基准（V1.6.18）──
        self.window_home = "900x650"  # 首开小窗口尺寸；从最大化还原须锁回此尺寸
        self.root.geometry(self.window_home)
        self.root.minsize(700, 500)
        self._prev_state = "normal"
        self.root.bind("<Configure>", self._on_resize)
        
        # ── 应用图标（V1.8.5）：标题栏左上角显示；exe 文件图标由 PyInstaller --icon 提供 ──
        try:
            _icon_base = sys._MEIPASS if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
            _icon_path = os.path.join(_icon_base, THEME['icon'])
            if os.path.exists(_icon_path):
                self.root.iconbitmap(_icon_path)
        except Exception:
            pass
        
        # Data file path - save in same directory as exe
        if getattr(sys, 'frozen', False):
            base_path = os.path.dirname(sys.executable)
        else:
            base_path = os.path.dirname(os.path.abspath(__file__))
        
        self.data_file = os.path.join(base_path, "todo_data.json")
        
        # Priority options (Eisenhower Matrix) —— 主题化：档位名按套不同，经验/颜色按位置解耦
        self.priority_options = THEME['priority_options']
        # 颜色按「位置索引」映射，与档位名解耦（顺序：最高→最低）
        self.priority_colors = dict(zip(self.priority_options,
                                        ["#FFCCCC", "#FFF2CC", "#D9EAD3", "#C9DAF8"]))
        # 经验值按「位置索引」（顺序：最高20 → 最低5）
        self.WISDOM_BY_PRIORITY = dict(zip(self.priority_options, [20, 20, 10, 5]))

        # Task category options —— 主题化
        self.category_options = THEME['category_options']
        
        # Tag options (predefined tags for quick selection)
        self.predefined_tags = ["紧急", "重要", "待讨论", "进行中", "阻塞", "待审核"]
        
        # Font configurations
        self.font_title = ("Microsoft YaHei", 20, "bold")
        self.font_label = ("Microsoft YaHei", 13)
        self.font_entry = ("Microsoft YaHei", 13)
        self.font_button = ("Microsoft YaHei", 12, "bold")
        self.font_table_header = ("Microsoft YaHei", 12, "bold")
        self.font_table_content = ("Microsoft YaHei", 12)
        
        # Load data
        self.data = self.load_data()
        self._data_loaded = True
        # 打开即刷新「今日已完成 + 进行中」摘要，确保 daily_summary 反映真实数据
        self.sync_daily_summary()
        
        # Timer variables (compact timer in pending tab)
        self.timer_running = False
        self.timer_paused = False
        self.timer_seconds = 25 * 60  # Default: 25 minutes
        self.timer_default_seconds = 25 * 60
        self.timer_after_id = None
        self.timer_display_var = tk.StringVar(value="⏰ 25:00")
        
        # Sorting criteria
        self.sort_criteria = tk.StringVar(value=self.T['sort_default'])
        
        # Search query
        self.search_query = tk.StringVar()
        self.search_query.trace_add('write', lambda *args: self.refresh_display())
        
        # Game data (冒险者游戏系统)
        self.game_data = self.load_game_data()
        
        # Animation variables
        self.animation_label = None
        self.animation_after_id = None
        
        # Setup GUI
        self.setup_gui()
        
        # Refresh display
        self.refresh_display()
    
    def load_data(self):
        """Load data from JSON file with backup recovery"""
        if os.path.exists(self.data_file):
            try:
                with open(self.data_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    if "tasks" not in data:
                        data["tasks"] = []
                    if "completed" not in data:
                        data["completed"] = []

                    # 统一迁移：分类/优先级/标签/支线，避免重复循环与重复落盘
                    category_map = THEME['legacy_category_map']
                    priority_map = THEME['legacy_priority_map']
                    migrated = False
                    for task in data["tasks"] + data["completed"]:
                        old_cat = task.get("category", "")
                        if old_cat in category_map:
                            task["category"] = category_map[old_cat]
                            migrated = True
                        old_pri = task.get("priority", "")
                        if old_pri in priority_map:
                            task["priority"] = priority_map[old_pri]
                            migrated = True
                        if "tags" not in task:
                            task["tags"] = []
                            migrated = True
                        if "subtasks" not in task:
                            task["subtasks"] = []
                            migrated = True

                    # 回补 wisdom_gain（V1.6.1 之前完成的任务未记录该字段，
                    # 会导致删除/恢复时已通关任务无法正确扣回本周智慧）— V1.6.2
                    for task in data["completed"]:
                        if "wisdom_gain" not in task:
                            task["wisdom_gain"] = self.WISDOM_BY_PRIORITY.get(task.get("priority", ""), 5)

                    # 仅当确有字段迁移/补值时统一落盘一次（避免无变化时也写文件）
                    if migrated:
                        try:
                            if os.path.exists(self.data_file):
                                shutil.copy2(self.data_file, self.data_file + '.bak')
                            with open(self.data_file, 'w', encoding='utf-8') as f:
                                json.dump(data, f, ensure_ascii=False, indent=2)
                        except Exception:
                            pass

                    return data
            except Exception as e:
                print(f"Error loading data: {e}")
                # Try to recover from backup
                backup_file = self.data_file + '.bak'
                if os.path.exists(backup_file):
                    try:
                        with open(backup_file, 'r', encoding='utf-8') as f:
                            data = json.load(f)
                            messagebox.showwarning("数据恢复", 
                                f"主数据文件损坏，已从备份恢复。\n错误: {str(e)}")
                            return data
                    except Exception as be:
                        print(f"Backup also corrupted: {be}")
                messagebox.showerror("错误", f"加载数据失败:\n{str(e)}\n将创建新的数据文件。")
        return {"tasks": [], "completed": []}
    
    def load_game_data(self):
        """Load game data (level, exp, attributes)."""
        # Game data is stored in the main data file under "game" key
        game_file = self.data_file.replace("todo_data.json", "game_data.json")

        # 当前周一（用于跨周判定与初始化）— V1.6.2 提到函数顶部，确保无文件分支也能正确初始化
        today = date.today()
        monday = today - timedelta(days=today.weekday())  # 本周一
        monday_str = monday.isoformat()

        # Default game data - V1.6.0：每周智慧系统
        default_game_data = {
            # 历史总智慧（保留，用于统计）
            "total_wisdom": 0,

            # 每周智慧系统
            "weekly_wisdom": 0,  # 本周获得的智慧
            "weekly_level": 1,  # 本周等级（1-6）
            "week_start_date": monday_str,  # 本周开始日期（周一的日期），默认初始化为当前周一
            "weekly_history": [],  # 历史周报记录

            # 统计信息
            "stats": {
                "total_completed": 0,
                "total_added": 0,
                "current_streak": 0,
                "longest_streak": 0
            }
        }

        if os.path.exists(game_file):
            try:
                with open(game_file, 'r', encoding='utf-8') as f:
                    game_data = json.load(f)

                    # ── 数据迁移（V1.6.0：从旧版本升级）──
                    if "exp" in game_data:  # 旧版本标志：存在 exp 即需迁移（不依赖 total_wisdom 是否缺失）
                        # 将旧的 exp 迁移到 total_wisdom
                        game_data["total_wisdom"] = game_data.get("exp", 0)

                        # 移除旧字段
                        old_keys = ["exp", "exp_to_next_level", "gold", "daily_exp", "daily_gold", "last_daily_date", "attributes"]
                        for old_key in old_keys:
                            if old_key in game_data:
                                del game_data[old_key]

                    # Ensure all fields exist
                    for key in default_game_data:
                        if key not in game_data:
                            game_data[key] = default_game_data[key]

                    # ── 检查是否需要重置每周统计 ──
                    if game_data.get("week_start_date") != monday_str:
                        # 跨周了，需要重置

                        # 保存上周数据到历史记录
                        if game_data.get("weekly_wisdom", 0) > 0:
                            # 计算上周的等级和称号
                            last_week_wisdom = game_data["weekly_wisdom"]
                            last_week_level = game_data["weekly_level"]
                            last_week_title = self.get_wisdom_title(last_week_level)

                            # 添加到历史记录
                            game_data["weekly_history"].append({
                                "week": game_data.get("week_start_date", ""),
                                "wisdom": last_week_wisdom,
                                "level": last_week_level,
                                "title": last_week_title
                            })

                        # 重置本周数据
                        game_data["weekly_wisdom"] = 0
                        game_data["weekly_level"] = 1
                        game_data["week_start_date"] = monday_str

                    return game_data
            except Exception as e:
                print(f"Error loading game data: {e}")
        
        return default_game_data
    
    def save_game_data(self):
        """Save game data to file."""
        try:
            game_file = self.data_file.replace("todo_data.json", "game_data.json")
            with open(game_file, 'w', encoding='utf-8') as f:
                json.dump(self.game_data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Error saving game data: {e}")
    
    @staticmethod
    def get_wisdom_title(level):
        """Calculate wisdom title based on LEVEL (主题化：读 THEME['levels'])."""
        levels = THEME['levels']
        idx = max(1, min(int(level), len(levels))) - 1
        return levels[idx]['name']

    @staticmethod
    def get_wisdom_for_level(level):
        """Return wisdom needed for each level (主题化：读 THEME['levels'])."""
        levels = THEME['levels']
        idx = max(1, min(int(level), len(levels))) - 1
        return (levels[idx]['min'], levels[idx]['max'])

    @staticmethod
    def get_level_from_wisdom(wisdom):
        """Calculate level from wisdom points (主题化：读 THEME['levels'])."""
        for i, lv in enumerate(THEME['levels'], 1):
            if lv['min'] <= wisdom <= lv['max']:
                return i
        return len(THEME['levels'])
    
    def _check_wisdom_level_up(self):
        """Check and process level up (V1.6.0)."""
        current_wisdom = self.game_data["weekly_wisdom"]
        current_level = self.game_data["weekly_level"]
        
        new_level = self.get_level_from_wisdom(current_wisdom)
        
        if new_level > current_level:
            # Level up!
            self.game_data["weekly_level"] = new_level
            self.show_level_up_animation()
        elif new_level < current_level:
            # Level down (should not happen, but for safety)
            self.game_data["weekly_level"] = new_level
    
    def add_wisdom(self, amount):
        """Add wisdom points (V1.6.0：每周智慧系统)."""
        self.game_data["weekly_wisdom"] += amount
        self.game_data["total_wisdom"] += amount
        
        # Check for level up
        self._check_wisdom_level_up()
        
        self.save_game_data()
        self.update_game_display()
    
    def subtract_wisdom(self, total_amount, weekly_amount):
        """Deduct wisdom points (V1.6.1：删除/恢复任务同步扣回).
        total_amount 始终扣回历史总智慧；weekly_amount 仅当任务属本周时>0。"""
        self.apply_wisdom_change(self.game_data, -total_amount, -weekly_amount)
        self.save_game_data()
        self.update_game_display()
    
    @staticmethod
    def apply_wisdom_change(game_data, total_delta, weekly_delta):
        """Apply a wisdom change to game_data (no save/refresh).
        Deltas may be negative for deduction; clamped at 0."""
        game_data["total_wisdom"] = max(0, game_data["total_wisdom"] + total_delta)
        game_data["weekly_wisdom"] = max(0, game_data["weekly_wisdom"] + weekly_delta)
        game_data["weekly_level"] = TodoApp.get_level_from_wisdom(game_data["weekly_wisdom"])
    
    @staticmethod
    def is_task_this_week(completed_str, week_start_date):
        """Return True if task completed at `completed_str` belongs to the week of `week_start_date`."""
        if not completed_str or not week_start_date:
            return False
        try:
            cd = datetime.strptime(completed_str, "%Y-%m-%d %H:%M:%S")
            cm = (cd - timedelta(days=cd.weekday())).strftime("%Y-%m-%d")
            return cm == week_start_date
        except Exception:
            return False
    
    def _apply_emoji_images(self, widget):
        """彩色 emoji 图片化管线已于 V1.8.9 移除（韦老板确认灰版观感更稳重），
        保留为空 no-op 以维持既有调用点兼容。"""
        pass

    def _set_emoji_text(self, widget, msg):
        """动态更新文本（灰版：纯文本，emoji 由 Tk 原生渲染）。"""
        try:
            widget.configure(text=msg)
        except Exception:
            pass

    def set_status(self, msg):
        """状态栏：左侧图标占位 + 右侧文本（灰版纯文本）。"""
        try:
            self.status_icon.configure(image="")
            self.status_text.configure(text=msg)
        except Exception:
            pass

    def show_wisdom_animation(self, amount):
        """Show wisdom gain animation (V1.6.0)."""
        if self.animation_label:
            self.animation_label.destroy()
        
        self.animation_label = tk.Label(
            self.root,
            text=self.T['wisdom_anim'].format(amount),
            font=("Microsoft YaHei", 24, "bold"),
            fg="#9B59B6",
            bg="#2B2B2B"
        )
        self.animation_label.place(relx=0.5, rely=0.3, anchor=tk.CENTER)
        self._apply_emoji_images(self.animation_label)
        
        # Animate: fade out and move up
        self._animate_exp(0)
    
    def _animate_exp(self, step):
        """Animate exp label (fade out and move up)."""
        if not self.animation_label:
            return
        
        if step >= 30:
            self.animation_label.destroy()
            self.animation_label = None
            return
        
        # Move up
        current_y = self.animation_label.winfo_y()
        self.animation_label.place(y=current_y - 3)
        
        # Fade effect (simulate by changing foreground color)
        alpha = int(255 * (1 - step / 30))
        color = f"#{alpha:02X}{alpha:02X}00"
        try:
            self.animation_label.config(fg=color)
        except:
            pass
        
        self.animation_after_id = self.root.after(30, lambda: self._animate_exp(step + 1))
    
    def show_level_up_animation(self):
        """Show level up animation (V1.6.0：每周智慧系统)."""
        level_up_window = tk.Toplevel(self.root)
        level_up_window.title(self.T['levelup_title'])
        level_up_window.geometry("380x300")
        level_up_window.transient(self.root)
        level_up_window.grab_set()

        # Center the window
        level_up_window.update_idletasks()
        x = (level_up_window.winfo_screenwidth() // 2) - (380 // 2)
        y = (level_up_window.winfo_screenheight() // 2) - (300 // 2)
        level_up_window.geometry(f"380x300+{x}+{y}")
        
        level = self.game_data.get('weekly_level', 1)
        title = self.get_wisdom_title(level)
        wisdom = self.game_data.get('weekly_wisdom', 0)
        
        tk.Label(level_up_window, text=self.T['levelup_icon'], font=("Microsoft YaHei", 48)).pack(pady=15)
        tk.Label(level_up_window, text=self.T['levelup_msg'], font=("Microsoft YaHei", 16, "bold")).pack()
        tk.Label(level_up_window, text=f"{title} | Lv.{level}", font=("Microsoft YaHei", 14, "bold"), fg=THEME['colors']['accent']).pack(pady=8)
        tk.Label(level_up_window, text=f"本周智慧: {wisdom}", font=("Microsoft YaHei", 12)).pack()
        
        tk.Button(level_up_window, text=self.T['levelup_btn'], command=level_up_window.destroy,
                 font=("Microsoft YaHei", 12), bg=THEME['colors']['primary'], fg="white").pack(pady=15)
        self._apply_emoji_images(level_up_window)
    
    def show_new_task_animation(self):
        """Show new task added animation."""
        if self.animation_label:
            self.animation_label.destroy()
        
        self.animation_label = tk.Label(
            self.root,
            text=self.T['msg_new_task_gen'],
            font=("Microsoft YaHei", 20, "bold"),
            fg="#00AA00",
            bg=THEME['colors']['panel_bg']
        )
        self.animation_label.place(relx=0.5, rely=0.4, anchor=tk.CENTER)
        self._apply_emoji_images(self.animation_label)
        
        # Animate: fade out
        self._animate_new_task(0)
    
    def _animate_new_task(self, step):
        """Animate new task label (fade out)."""
        if not self.animation_label:
            return
        
        if step >= 25:
            self.animation_label.destroy()
            self.animation_label = None
            return
        
        # Fade effect
        alpha = int(255 * (1 - step / 25))
        color = f"#00{alpha:02X}00"
        try:
            self.animation_label.config(fg=color)
        except:
            pass
        
        self.animation_after_id = self.root.after(30, lambda: self._animate_new_task(step + 1))
    
    def show_fireworks_animation(self):
        """完成任务庆祝：居中庆祝卡片 + 主题色礼花（美化版，V2.0.3）。"""
        # 防重入：销毁上一次未结束的庆祝窗
        prev = getattr(self, '_celebrate_win', None)
        if prev and prev.winfo_exists():
            try:
                prev.destroy()
            except Exception:
                pass

        win = tk.Toplevel(self.root)
        self._celebrate_win = win
        win.overrideredirect(True)
        win.attributes('-topmost', True)
        TRANS = '#abcdef'          # 透明色键（卡片区域外透出主界面）
        win.configure(bg=TRANS)
        try:
            win.attributes('-transparentcolor', TRANS)
        except Exception:
            pass
        win.attributes('-alpha', 0.0)

        W, H = 480, 380
        self.root.update_idletasks()
        rx, ry = self.root.winfo_rootx(), self.root.winfo_rooty()
        rw, rh = self.root.winfo_width(), self.root.winfo_height()
        x = rx + max(0, (rw - W) // 2)
        y = ry + max(0, (rh - H) // 2)
        win.geometry(f"{W}x{H}+{x}+{y}")

        accent = self.T.get('celebrate_accent', '#F1C40F')
        navy = self.T.get('title_color', '#2C3E50')

        canvas = tk.Canvas(win, width=W, height=H, bg=TRANS, highlightthickness=0)
        canvas.pack(fill='both', expand=True)

        # 礼花粒子（先画，置于卡片之下，呈现从卡片后迸发的层次）
        ccx, ccy = W // 2, H // 2
        conf_palette = [accent, navy, '#5B8FF9', '#E67E22', '#2ECC71']
        particles = []
        for _ in range(80):
            a = random.uniform(0, math.pi * 2)
            sp = random.uniform(3, 8)
            size = random.randint(4, 9)
            color = random.choice(conf_palette)
            pid = canvas.create_oval(ccx - size, ccy - size, ccx + size, ccy + size,
                                     fill=color, outline=color)
            particles.append({'id': pid, 'x': ccx, 'y': ccy,
                              'vx': math.cos(a) * sp, 'vy': math.sin(a) * sp - 2,
                              'size': size, 'life': 70})

        # 卡片：阴影 + 圆角白卡（画在礼花之上，盖住中心）
        cx, cy, cw, ch = 60, 70, 360, 240
        self._round_rect(canvas, cx, cy + 6, cw, ch, 18, fill='#d9dee4', outline='#d9dee4')
        self._round_rect(canvas, cx, cy, cw, ch, 18, fill='#ffffff', outline='#e6ebf0')

        emoji = self.T.get('complete_emoji', '🎉')
        title = self.T.get('complete_title', '恭喜您，冒险已完成！')
        sub = self.T.get('complete_sub', '本次冒险已记入战绩')
        chip = self.T.get('complete_chip', '+10 经验值')
        chip_fg = self.T.get('chip_fg', '#2C3E50')
        chip_bg = self.T.get('chip_bg', '#F1C40F')

        tk.Label(win, text=emoji, font=('Microsoft YaHei', 46), bg='#ffffff').place(relx=0.5, y=cy + 46, anchor='center')
        tk.Label(win, text=title, font=('Microsoft YaHei', 17, 'bold'), fg=navy, bg='#ffffff').place(relx=0.5, y=cy + 96, anchor='center')
        tk.Label(win, text=sub, font=('Microsoft YaHei', 12), fg='#8a97a0', bg='#ffffff').place(relx=0.5, y=cy + 140, anchor='center')

        # 经验值药丸：canvas 圆角底衬 + 透明底 Label 文字
        chip_label = tk.Label(win, text=chip, font=('Microsoft YaHei', 13, 'bold'),
                              fg=chip_fg, bg=TRANS, padx=18, pady=7)
        chip_label.place(relx=0.5, y=cy + 206, anchor='center')
        win.update_idletasks()
        bx = chip_label.winfo_rootx() - win.winfo_rootx()
        by = chip_label.winfo_rooty() - win.winfo_rooty()
        bw = chip_label.winfo_width()
        bh = chip_label.winfo_height()
        self._round_rect(canvas, bx - 2, by - 2, bw + 4, bh + 4, 10, fill=chip_bg, outline=chip_bg)

        # 入场淡入 + 光环 + 礼花动画
        self._celebrate_fade(win, 1.0, 1)
        self._celebrate_ring(canvas, ccx, ccy, 0, accent)
        self._celebrate_confetti(canvas, particles, 0)

        # 约 2.4s 后淡出销毁
        win.after(2400, lambda: self._celebrate_fade(win, 0.0, -1, done=lambda: self._safe_destroy(win)))
    
    def _round_rect(self, c, x, y, w, h, r, fill, outline):
        """在 Canvas 上画圆角矩形（四角圆弧拼法）。"""
        c.create_arc(x, y, x + 2*r, y + 2*r, start=90, extent=90, style='pieslice', fill=fill, outline=outline)
        c.create_arc(x + w - 2*r, y, x + w, y + 2*r, start=0, extent=90, style='pieslice', fill=fill, outline=outline)
        c.create_arc(x, y + h - 2*r, x + 2*r, y + h, start=180, extent=90, style='pieslice', fill=fill, outline=outline)
        c.create_arc(x + w - 2*r, y + h - 2*r, x + w, y + h, start=270, extent=90, style='pieslice', fill=fill, outline=outline)
        c.create_rectangle(x + r, y, x + w - r, y + h, fill=fill, outline=outline)
        c.create_rectangle(x, y + r, x + w, y + h - r, fill=fill, outline=outline)

    def _safe_destroy(self, win):
        try:
            win.destroy()
        except Exception:
            pass

    def _celebrate_fade(self, win, target, step, done=None):
        """淡入/淡出整个庆祝窗（step>0 淡入，step<0 淡出）。"""
        try:
            cur = win.attributes('-alpha')
        except Exception:
            cur = 1.0
        if step > 0 and cur < target:
            cur = min(target, cur + 0.08)
            try:
                win.attributes('-alpha', cur)
            except Exception:
                pass
            win.after(30, lambda: self._celebrate_fade(win, target, step, done))
        elif step < 0 and cur > target:
            cur = max(target, cur - 0.08)
            try:
                win.attributes('-alpha', cur)
            except Exception:
                pass
            win.after(30, lambda: self._celebrate_fade(win, target, step, done))
        else:
            if done:
                done()

    def _celebrate_confetti(self, canvas, particles, step):
        if step > 65:
            return
        for p in particles:
            if p['life'] <= 0:
                continue
            p['vy'] += 0.12
            p['x'] += p['vx']
            p['y'] += p['vy']
            p['life'] -= 1
            try:
                canvas.coords(p['id'], p['x'] - p['size'], p['y'] - p['size'],
                              p['x'] + p['size'], p['y'] + p['size'])
            except Exception:
                pass
        canvas.after(30, lambda: self._celebrate_confetti(canvas, particles, step + 1))

    def _celebrate_ring(self, canvas, cx, cy, step, accent):
        if step > 18:
            return
        r = 18 + step * 2
        try:
            canvas.delete('cring')
            canvas.create_oval(cx - r, cy - r, cx + r, cy + r,
                               outline=accent, width=3, tags='cring')
        except Exception:
            pass
        canvas.after(25, lambda: self._celebrate_ring(canvas, cx, cy, step + 1, accent))
    
    def update_game_display(self):
        """Update growth system panels in stats tab (V1.6.0：每周智慧系统)."""
        try:
            # 更新本周智慧面板
            weekly_wisdom = self.game_data.get('weekly_wisdom', 0)
            weekly_level = self.game_data.get('weekly_level', 1)
            week_start_date = self.game_data.get('week_start_date', '')
            
            # 称号由等级决定
            wisdom_title = self.get_wisdom_title(weekly_level)
            
            # 进度由当前智慧和下一级所需智慧决定
            min_wisdom, max_wisdom = self.get_wisdom_for_level(weekly_level)
            if weekly_level >= 6:
                # 已满级
                wisdom_progress = 1.0
                next_title = "已满级"
                next_level = 6
            else:
                wisdom_progress = (weekly_wisdom - min_wisdom) / (max_wisdom - min_wisdom) if max_wisdom > min_wisdom else 0.0
                next_level = weekly_level + 1
                next_title = self.get_wisdom_title(next_level)
            
            if hasattr(self, 'exp_rank_label'):
                self.exp_rank_label.config(text=f"{wisdom_title} | Lv.{weekly_level}")
            if hasattr(self, 'exp_daily_label'):
                # 显示本周剩余天数
                today = date.today()
                monday = today - timedelta(days=today.weekday())
                days_passed = (today - monday).days + 1
                days_left = 7 - days_passed
                self._set_emoji_text(self.exp_daily_label, self.T['exp_daily'].format(days_passed, days_left))
            if hasattr(self, 'exp_detail_label'):
                if weekly_level >= 6:
                    self.exp_detail_label.config(text=f"{weekly_wisdom} 智慧 (已满级)")
                else:
                    self.exp_detail_label.config(text=f"{weekly_wisdom}/{max_wisdom} 智慧 → {next_title} (Lv.{next_level})")
            if hasattr(self, 'exp_progress_canvas'):
                self._draw_progress_bar(self.exp_progress_canvas, wisdom_progress, THEME['colors']['accent'])
            
            # 金币面板已移除（V1.6.0）
                    
        except Exception as e:
            print(f"Error updating game display: {e}")
    
    def _draw_progress_bar(self, canvas, progress, color):
        """在Canvas上绘制进度条"""
        try:
            canvas.delete("all")
            w = canvas.winfo_width()
            h = canvas.winfo_height()
            if w <= 1:
                w = 320
            if h <= 1:
                h = 24
            
            # 背景
            canvas.create_rectangle(0, 0, w, h, fill=THEME['colors']['dark_alt'] if color == THEME['colors']['accent'] else THEME['colors']['dark_alt2'], outline="")
            
            # 进度
            pw = int(w * max(0.0, min(1.0, progress)))
            if pw > 0:
                canvas.create_rectangle(0, 0, pw, h, fill=color, outline="")
            
            # 百分比文字
            pct = int(progress * 100)
            canvas.create_text(w // 2, h // 2, text=f"{pct}%", fill="white", font=("Microsoft YaHei", 10, "bold"))
        except:
            pass
    
    def save_data(self):
        """Save data to JSON file with backup"""
        try:
            # Create backup before saving
            if os.path.exists(self.data_file):
                backup_file = self.data_file + '.bak'
                shutil.copy2(self.data_file, backup_file)
            
            with open(self.data_file, 'w', encoding='utf-8') as f:
                json.dump(self.data, f, ensure_ascii=False, indent=2)
            # 仅当数据成功落盘后再同步摘要，避免保存失败却写出（可能不完整的）摘要
            self.sync_daily_summary()
        except Exception as e:
            print(f"Error saving data: {e}")
            messagebox.showerror("错误", f"保存数据失败:\n{str(e)}")
    
    def sync_daily_summary(self):
        """将「今日已完成 + 进行中」同步写入 daily_summary.txt（UTF-8 无 BOM），供快速汇报省 token。
        在 save_data 末尾自动调用；只写不读，异常仅打印不影响主程序。
        今日已完成 = completed 字段日期为今天的任务（含其下已完成支线）；
        进行中 = 当前 pending 任务（含其下未完成支线）。"""
        if not getattr(self, '_data_loaded', False):
            return
        try:
            summary_path = os.path.join(os.path.dirname(self.data_file), 'daily_summary.txt')
            today = datetime.now().strftime('%Y-%m-%d')

            done_lines = []
            for t in self.data.get('completed', []):
                if t.get('completed', '')[:10] == today:
                    done_lines.append(f" - {t.get('text', '')}")
                    for st in t.get('subtasks', []):
                        if st.get('done'):
                            done_lines.append(f"   - [支线] {st.get('text', '')}")

            doing_lines = []
            for t in self.data.get('tasks', []):
                doing_lines.append(f" - {t.get('text', '')}")
                for st in t.get('subtasks', []):
                    if not st.get('done'):
                        doing_lines.append(f"   - [支线] {st.get('text', '')}")

            content = "今日已完成:\n"
            content += ("\n".join(done_lines) + "\n") if done_lines else "（今日暂无）\n"
            content += "进行中:\n"
            content += ("\n".join(doing_lines) + "\n") if doing_lines else "（暂无）\n"

            with open(summary_path, 'w', encoding='utf-8') as f:
                f.write(content)
        except Exception as e:
            print(f"sync_daily_summary skipped: {e}")
    
    # ──────────────────────────────────────────────
    #  INPUT AREA — using pack() nested frames
    #  This reliably shows all widgets regardless
    #  of window size.
    # ──────────────────────────────────────────────
    def _build_input_area(self):
        """Build the add-task area - same row layout."""
        input_frame = tk.LabelFrame(self.root, text="",
                                    font=self.font_label,
                                    padx=12, pady=10)
        input_frame.pack(fill=tk.X, padx=15, pady=(10, 5))

        # ── Row 0: label + entry + button (SAME ROW) ──
        row0 = tk.Frame(input_frame)
        row0.pack(fill=tk.X, pady=(0, 8))

        tk.Label(row0, text=self.T['lbl_desc_add'], font=self.font_label).pack(side=tk.LEFT, padx=(0, 8))

        self.task_entry = tk.Entry(row0, font=self.font_entry)
        self.task_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 12))
        self.task_entry.bind('<Return>', lambda e: self.add_task())

        add_btn = tk.Button(row0, text=self.T['btn_add_task'], command=self.add_task,
                            bg=THEME['colors']['primary'], fg="white", font=self.font_button,
                            cursor="hand2", padx=20, pady=4)
        add_btn.pack(side=tk.RIGHT)

        # ── Row 1: priority / category / person ──
        row1 = tk.Frame(input_frame)
        row1.pack(fill=tk.X)

        # -- Priority group --
        g1 = tk.Frame(row1)
        g1.pack(side=tk.LEFT, padx=(0, 18))
        tk.Label(g1, text=self.T['lbl_priority'], font=self.font_label).pack(side=tk.LEFT, padx=(0, 6))
        self.priority_var = tk.StringVar(value=self.priority_options[0])
        self.priority_combo = ttk.Combobox(g1, textvariable=self.priority_var,
                                           values=self.priority_options,
                                           state='readonly', width=16, font=self.font_entry)
        self.priority_combo.pack(side=tk.LEFT)

        # -- Category group --
        g2 = tk.Frame(row1)
        g2.pack(side=tk.LEFT, padx=(0, 18))
        tk.Label(g2, text=self.T['lbl_type'], font=self.font_label).pack(side=tk.LEFT, padx=(0, 6))
        self.category_var = tk.StringVar(value=self.category_options[0])
        self.category_combo = ttk.Combobox(g2, textvariable=self.category_var,
                                           values=self.category_options,
                                           state='readonly', width=14, font=self.font_entry)
        self.category_combo.pack(side=tk.LEFT)

        # -- Person group --
        g3 = tk.Frame(row1)
        g3.pack(side=tk.LEFT, padx=(0, 18))
        tk.Label(g3, text=self.T['lbl_owner'], font=self.font_label).pack(side=tk.LEFT, padx=(0, 6))
        self.person_entry = tk.Entry(g3, width=10, font=self.font_entry)
        self.person_entry.pack(side=tk.LEFT)
        
        # -- Tags group --
        g4 = tk.Frame(row1)
        g4.pack(side=tk.LEFT)
        tk.Label(g4, text=self.T['form_tag'], font=self.font_label).pack(side=tk.LEFT, padx=(0, 6))
        self.tags_entry = tk.Entry(g4, width=15, font=self.font_entry)
        self.tags_entry.pack(side=tk.LEFT)
        tk.Label(g4, text="(逗号分隔)", font=("Microsoft YaHei", 10), fg="#888888").pack(side=tk.LEFT, padx=(4, 0))

        return input_frame

    # ──────────────────────────────────────────────
    #  MAIN GUI
    # ──────────────────────────────────────────────
    def setup_gui(self):
        """Setup the GUI components"""
        # Apply ttk theme for better aesthetics
        style = ttk.Style()
        available_themes = style.theme_names()
        if 'clam' in available_themes:
            style.theme_use('clam')
        # Customize theme colors
        style.configure('TNotebook.Tab', padding=[12, 8], font=self.font_label)
        style.configure('Treeview', font=self.font_table_content, rowheight=45)
        style.configure('Treeview.Heading', font=self.font_table_header)
        
        # Title bar
        title_frame = tk.Frame(self.root, bg=THEME['colors']['title_bar'], height=60)
        title_frame.pack(fill=tk.X)
        title_frame.pack_propagate(False)

        # 居中内层组：图标 + 文字（紧凑排列，间距 padx=3）
        title_center = tk.Frame(title_frame, bg=THEME['colors']['title_bar'])
        title_center.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

        title_icon_file = THEME.get('title_icon')
        if title_icon_file:
            try:
                _ti_base = sys._MEIPASS if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
                _ti_path = os.path.join(_ti_base, title_icon_file)
                if os.path.exists(_ti_path):
                    self._title_icon_img = tk.PhotoImage(file=_ti_path)
                    tk.Label(title_center, image=self._title_icon_img,
                             bg=THEME['colors']['title_bar']).pack(side=tk.LEFT, padx=(0, 3))
            except Exception:
                self._title_icon_img = None  # 加载失败回退纯文字

        title_label = tk.Label(title_center, text=THEME['title'],
                               font=self.font_title,
                               bg=THEME['colors']['title_bar'], fg="white")
        title_label.pack(side=tk.LEFT)

        # Input area (pack layout — reliable)
        self._build_input_area()

        # Notebook (tabs) with custom tab style
        style = ttk.Style()
        style.configure("TNotebook.Tab", padding=[12, 8], font=self.font_label)
        style.map("TNotebook.Tab",
            background=[("selected", THEME['colors']['panel_bg'])])  # Light gray when selected/focused
        
        notebook = ttk.Notebook(self.root)
        self.notebook = notebook
        notebook.pack(fill=tk.BOTH, expand=True, padx=15, pady=(5, 10))

        pending_frame = tk.Frame(notebook)
        notebook.add(pending_frame, text=self.T['tab_pending'])
        self.setup_pending_tab(pending_frame)

        completed_frame = tk.Frame(notebook)
        notebook.add(completed_frame, text=self.T['tab_completed'])
        self.setup_completed_tab(completed_frame)

        stats_frame = tk.Frame(notebook)
        notebook.add(stats_frame, text=self.T['tab_stats'])
        self.setup_stats_tab(stats_frame)
        
        # Calendar tab
        
        # ── 底部状态栏（灰版：图标占位 + 文本）──
        self.status_frame = tk.Frame(self.root, bg=THEME['colors']['status_bar'])
        self.status_frame.pack(side=tk.BOTTOM, fill=tk.X)
        self.status_icon = tk.Label(self.status_frame, bg=THEME['colors']['status_bar'])
        self.status_icon.pack(side=tk.LEFT, padx=(10, 0))
        self.status_text = tk.Label(self.status_frame, text=self.T['status_ready'],
                                    font=("Microsoft YaHei", 10),
                                    bg=THEME['colors']['status_bar'], fg="#333333",
                                    anchor=tk.W, padx=2, pady=4)
        self.status_text.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.set_status(self.T['status_ready_icon'])

        # Update game display initially
        self.update_game_display()
        # 方案A：把界面 emoji 渲染为彩色图片（绕过 Tkinter 单色字形）
        self._apply_emoji_images(self.root)

    def setup_pending_tab(self, parent):
        """Setup pending tasks tab"""
        toolbar = tk.Frame(parent)
        toolbar.pack(fill=tk.X, pady=(10, 10))
        
        tk.Button(toolbar, text=self.T['btn_mark'], command=self.complete_task,
                  bg="#5CB85C", fg="white", font=self.font_button,
                  cursor="hand2", padx=16, pady=4).pack(side=tk.LEFT, padx=(0, 8))
        
        tk.Button(toolbar, text=self.T['btn_abandon'], command=self.delete_task,
                  bg="#D9534F", fg="white", font=self.font_button,
                  cursor="hand2", padx=16, pady=4).pack(side=tk.LEFT, padx=(0, 8))

        tk.Button(toolbar, text=self.T['btn_edit'], command=self.edit_task,
                  bg="#F0AD4E", fg="white", font=self.font_button,
                  cursor="hand2", padx=16, pady=4).pack(side=tk.LEFT, padx=(0, 8))

        tk.Button(toolbar, text=self.T['btn_refresh'], command=self.reload_data,
                  bg="#6C757D", fg="white", font=self.font_button,
                  cursor="hand2", padx=16, pady=4).pack(side=tk.LEFT, padx=(0, 8))
        
        # ── Sorting Dropdown ──
        sort_frame = tk.Frame(toolbar)
        sort_frame.pack(side=tk.LEFT, padx=(20, 0))
        
        tk.Label(sort_frame, text="排序：", font=self.font_label).pack(side=tk.LEFT, padx=(0, 5))
        self.sort_combobox = ttk.Combobox(sort_frame, textvariable=self.sort_criteria,
                                             values=[self.T['sort_default'], self.T['sort_priority'], self.T['sort_created'], self.T['sort_desc']],
                                             state="readonly", width=12, font=self.font_label)
        self.sort_combobox.pack(side=tk.LEFT)
        self.sort_combobox.bind('<<ComboboxSelected>>', lambda e: self.refresh_display())
        
        # ── Search Box ──
        search_frame = tk.Frame(toolbar)
        search_frame.pack(side=tk.LEFT, padx=(20, 0))
        
        tk.Label(search_frame, text="🔍 搜索：", font=self.font_label).pack(side=tk.LEFT, padx=(0, 5))
        self.search_entry = ttk.Entry(search_frame, textvariable=self.search_query, width=15, font=self.font_label)
        self.search_entry.pack(side=tk.LEFT)
        
        # ── Compact Timer (right-aligned) ──
        timer_frame = tk.Frame(toolbar)
        timer_frame.pack(side=tk.RIGHT, padx=(0, 20))
        
        self.timer_label = tk.Label(timer_frame, textvariable=self.timer_display_var,
                                   font=("Microsoft YaHei", 16, "bold"),
                                   padx=8)
        self.timer_label.pack(side=tk.LEFT)
        
        self.timer_btn = tk.Button(timer_frame, text="▶ 开始", command=self.toggle_timer,
                                  bg="#3498DB", fg="white", font=("Microsoft YaHei", 11, "bold"),
                                  cursor="hand2", padx=12, pady=4)
        self.timer_btn.pack(side=tk.LEFT, padx=(8, 4))
        
        self.timer_reset_btn = tk.Button(timer_frame, text="⏹ 重置", command=self.reset_timer,
                                        bg="#95A5A6", fg="white", font=("Microsoft YaHei", 11, "bold"),
                                        cursor="hand2", padx=12, pady=4)
        self.timer_reset_btn.pack(side=tk.LEFT)
        
        # Right-click on timer label to change time
        self.timer_label.bind('<Button-3>', self.show_timer_menu)
        self.timer_menu = tk.Menu(self.root, tearoff=0, font=self.font_label)
        self.timer_menu.add_command(label="5分钟", command=lambda: self.set_timer(5))
        self.timer_menu.add_command(label="15分钟", command=lambda: self.set_timer(15))
        self.timer_menu.add_command(label="25分钟", command=lambda: self.set_timer(25))
        self.timer_menu.add_command(label="30分钟", command=lambda: self.set_timer(30))
        
        # Table
        table_frame = tk.Frame(parent)
        table_frame.pack(fill=tk.BOTH, expand=True)
        
        columns = ('☑️', '编号', self.T['col_desc'], self.T['col_type'], self.T['col_priority'], self.T['col_owner'], self.T['col_tag'], self.T['col_progress'], self.T['col_receive'])
        self.pending_tree = ttk.Treeview(table_frame, columns=columns, show='tree headings', height=20)
        
        style = ttk.Style()
        style.configure("Treeview", font=self.font_table_content, rowheight=45)
        style.configure("Treeview.Heading", font=self.font_table_header)
        
        # 红色警报 任务加粗字体
        self.pending_tree.tag_configure(self.priority_options[0], font=('Microsoft YaHei', 11, 'bold'))
        
        # 主线任务加粗（需求：进行中页主线加粗、支线不加粗；已通关页不动）
        self.pending_tree.tag_configure('main', font=('Microsoft YaHei', 12, 'bold'))

        # 树形层级列（最左折叠三角），支线挂到主线下由 Tk 自动缩进
        self.pending_tree.column('#0', width=26, minwidth=26, stretch=False)
        self.pending_tree.heading('#0', text='')
        
        col_widths = {
            '☑️': 35,
            '编号': 40,
            self.T['col_desc']: 300,
            self.T['col_type']: 100,
            self.T['col_priority']: 100,
            self.T['col_owner']: 100,
            self.T['col_tag']: 90,
            self.T['col_progress']: 70,
            self.T['col_receive']: 140
        }
        # 各列最小宽度保底：窗口缩小时列名不再被截断（方案A，V1.6.20）
        col_minwidths = {
            '☑️': 30,
            '编号': 50,
            self.T['col_desc']: 110,
            self.T['col_type']: 95,
            self.T['col_priority']: 95,
            self.T['col_owner']: 85,
            self.T['col_tag']: 75,
            self.T['col_progress']: 75,
            self.T['col_receive']: 150
        }
        for col in columns:
            self.pending_tree.heading(col, text=col, anchor=tk.CENTER)
            # 方案B（V1.6.20修正）：仅「冒险描述」stretch=True 吸收伸缩，其余列全部钉死
            # → 全屏后缩小 = 首开小窗口效果（核心列宽度不变）
            anchor = tk.W if col in [self.T['col_owner'], self.T['col_tag']] else tk.CENTER
            w = col_widths.get(col, 100)
            mw = col_minwidths.get(col, 40)
            if col == self.T['col_desc']:
                self.pending_tree.column(col, width=w, minwidth=mw, stretch=True, anchor=anchor)
            elif col != self.T['col_receive']:  # 领取时间单独在循环外处理
                self.pending_tree.column(col, stretch=False, width=w, minwidth=mw, anchor=anchor)
        # 时间列不参与拉伸压缩，窗口缩小时始终完整显示时间戳（V1.6.19）
        self.pending_tree.column(self.T['col_receive'], stretch=False, width=140, minwidth=140)
        
        vsb = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.pending_tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.pending_tree.xview)
        self.pending_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.pending_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
        
        # 绑定复选框点击事件
        self.pending_tree.bind('<Button-1>', self.on_checkbox_click)
        
        # 绑定右键菜单
        self.pending_tree.bind('<Button-3>', self.show_context_menu)
    
    def show_context_menu(self, event):
        """Show right-click context menu for pending tasks"""
        try:
            # Select the item under cursor
            item = self.pending_tree.identify_row(event.y)
            
            # Create context menu dynamically
            context_menu = tk.Menu(self.root, tearoff=0, font=self.font_label)
            
            if item:
                # Check if this is a subtask
                is_subtask = "_subtask_" in item
                
                if is_subtask:
                    # Selected a subtask - show menu with toggle/delete/edit
                    self.pending_tree.selection_set(item)
                    context_menu.add_command(label="✅ 完成支线任务", command=self.toggle_subtask_from_menu)
                    context_menu.add_command(label="✏️ 修改任务描述", command=self.edit_subtask_from_menu)
                    context_menu.add_command(label="🗑 放弃支线任务", command=self.delete_subtask_from_menu)
                    context_menu.add_separator()
                    context_menu.add_command(label=self.T['btn_refresh'], command=self.reload_data)
                else:
                    # Selected a main task - show menu with delete/edit/add subtask
                    self.pending_tree.selection_set(item)
                    context_menu.add_command(label=self.T['btn_abandon'], command=self.delete_task)
                    context_menu.add_command(label=self.T['btn_edit'], command=self.edit_task)
                    context_menu.add_command(label=self.T['ctx_add_subtask'], command=self.context_add_subtask)
                    context_menu.add_separator()
                    context_menu.add_command(label=self.T['btn_refresh'], command=self.reload_data)
            else:
                # No task selected - show menu with add/refresh
                context_menu.add_command(label=self.T['ctx_new_task'], command=self.show_add_task_dialog)
                context_menu.add_separator()
                context_menu.add_command(label=self.T['btn_refresh'], command=self.reload_data)
            
            context_menu.post(event.x_root, event.y_root)
        except Exception as e:
            print(f"Error showing context menu: {e}")
    
    def context_add_subtask(self):
        """Add a subtask directly from context menu."""
        try:
            selection = self.pending_tree.selection()
            if not selection:
                messagebox.showwarning("警告", self.T['warn_select_sub'])
                return
            
            task_id = int(selection[0])
            task = None
            for t in self.data["tasks"]:
                if t["id"] == task_id:
                    task = t
                    break
            
            if not task:
                messagebox.showerror("错误", self.T['err_not_found'])
                return
            
            # Directly show the add subtask dialog
            self.show_add_subtask_dialog(task)
        except Exception as e:
            print(f"Error adding subtask from context menu: {e}")
    
    def context_manage_subtasks(self):
        """Manage subtasks from context menu."""
        try:
            selection = self.pending_tree.selection()
            if not selection:
                messagebox.showwarning("警告", self.T['warn_select_sub_manage'])
                return
            
            task_id = int(selection[0])
            task = None
            for t in self.data["tasks"]:
                if t["id"] == task_id:
                    task = t
                    break
            
            if not task:
                messagebox.showerror("错误", self.T['err_not_found'])
                return
            
            self.manage_subtasks(task)
        except Exception as e:
            print(f"Error managing subtasks from context menu: {e}")
    
    def toggle_subtask_from_menu(self):
        """Toggle subtask completion from context menu."""
        try:
            selection = self.pending_tree.selection()
            if not selection:
                return
            
            item = selection[0]
            if "_subtask_" in item:
                # Parse subtask iid: "{task_id}_subtask_{subtask_idx}"
                parts = item.split("_subtask_")
                task_id = int(parts[0])
                subtask_idx = int(parts[1]) - 1  # Convert to 0-based index
                
                # Find the task and toggle subtask
                for task in self.data["tasks"]:
                    if task["id"] == task_id:
                        if "subtasks" in task and subtask_idx < len(task["subtasks"]):
                            # Toggle subtask done status
                            task["subtasks"][subtask_idx]["done"] = not task["subtasks"][subtask_idx].get("done", False)
                            self.save_data()
                            self.refresh_display()
                            self.set_status("✓ 支线任务状态已更新！")
                        break
        except Exception as e:
            print(f"Error toggling subtask: {e}")
    
    def delete_subtask_from_menu(self):
        """Delete subtask from context menu."""
        try:
            selection = self.pending_tree.selection()
            if not selection:
                return
            
            item = selection[0]
            if "_subtask_" in item:
                # Parse subtask iid: "{task_id}_subtask_{subtask_idx}"
                parts = item.split("_subtask_")
                task_id = int(parts[0])
                subtask_idx = int(parts[1]) - 1  # Convert to 0-based index
                
                # Confirm deletion
                if not messagebox.askyesno("确认删除", "确定要删除这个支线任务吗？"):
                    return
                
                # Find the task and delete subtask
                for task in self.data["tasks"]:
                    if task["id"] == task_id:
                        if "subtasks" in task and subtask_idx < len(task["subtasks"]):
                            # Delete subtask
                            task["subtasks"].pop(subtask_idx)
                            self.save_data()
                            self.refresh_display()
                            self.set_status("🗑 支线任务已删除！")
                        break
        except Exception as e:
            print(f"Error deleting subtask: {e}")
    
    def edit_subtask_from_menu(self):
        """Edit subtask text from context menu."""
        try:
            selection = self.pending_tree.selection()
            if not selection:
                return
            
            item = selection[0]
            if "_subtask_" not in item:
                return
            
            # Parse subtask iid: "{task_id}_subtask_{subtask_idx}"
            parts = item.split("_subtask_")
            task_id = int(parts[0])
            subtask_idx = int(parts[1]) - 1  # Convert to 0-based index
            
            # Find the task and subtask
            task = None
            subtask = None
            for t in self.data["tasks"]:
                if t["id"] == task_id:
                    task = t
                    if "subtasks" in t and subtask_idx < len(t["subtasks"]):
                        subtask = t["subtasks"][subtask_idx]
                    break
            
            if not subtask:
                messagebox.showerror("错误", "未找到该支线任务！")
                return
            
            # Create edit dialog
            dialog = tk.Toplevel(self.root)
            dialog.title(f"✏️ 修改任务描述")
            dialog.geometry("400x210")
            dialog.resizable(False, False)
            dialog.transient(self.root)
            dialog.grab_set()
            
            # Center the dialog
            dialog.update_idletasks()
            width = dialog.winfo_width()
            height = dialog.winfo_height()
            x = (dialog.winfo_screenwidth() // 2) - (width // 2)
            y = (dialog.winfo_screenheight() // 2) - (height // 2)
            dialog.geometry(f"{width}x{height}+{x}+{y}")
            
            # Label
            tk.Label(dialog, text="📝 支线任务描述:", font=self.font_label).pack(anchor=tk.W, padx=20, pady=(20, 5))
            
            # Entry
            text_var = tk.StringVar(value=subtask.get("text", ""))
            entry = tk.Entry(dialog, font=self.font_entry, textvariable=text_var, width=40)
            entry.pack(padx=20, pady=(0, 12))
            entry.focus_set()
            entry.select_range(0, tk.END)
            
            # 标签
            tk.Label(dialog, text=self.T['dlg_tag'], font=self.font_label).pack(anchor=tk.W, padx=20, pady=(0, 4))
            tags_var = tk.StringVar(value=", ".join(subtask.get("tags", [])))
            tags_entry = tk.Entry(dialog, font=self.font_entry, textvariable=tags_var, width=40)
            tags_entry.pack(padx=20, pady=(0, 20))
            
            # Buttons
            btn_frame = tk.Frame(dialog)
            btn_frame.pack(pady=(0, 20))
            
            def save_edit():
                new_text = text_var.get().strip()
                if not new_text:
                    messagebox.showwarning("警告", "支线任务描述不能为空！")
                    return
                
                # 校验并更新标签
                tags_text = tags_var.get().strip()
                is_valid, error_msg, tags = self.validate_tags(tags_text)
                if not is_valid:
                    messagebox.showwarning("警告", error_msg)
                    return
                
                # Update subtask text
                subtask["text"] = new_text
                subtask["tags"] = tags
                self.save_data()
                self.refresh_display()
                self.set_status("✏️ 支线任务已更新！")
                dialog.destroy()
            
            tk.Button(btn_frame, text="取消", command=dialog.destroy,
                     font=self.font_button, padx=20, pady=5).pack(side=tk.LEFT, padx=(0, 10))
            tk.Button(btn_frame, text="✅ 保存", command=save_edit,
                     bg=THEME['colors']['primary'], fg="white", font=self.font_button,
                     cursor="hand2", padx=20, pady=5).pack(side=tk.LEFT)
            
            # Bind Enter key to save
            dialog.bind('<Return>', lambda e: save_edit())
            self._apply_emoji_images(dialog)
            
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            messagebox.showerror("编辑支线任务出错", f"错误详情:\n{str(e)}")
    
    def setup_completed_tab(self, parent):
        """Setup completed tasks tab"""
        # Filter
        filter_frame = tk.Frame(parent)
        filter_frame.pack(fill=tk.X, pady=(10, 10))

        tk.Label(filter_frame, text="🔍 筛选日期:", font=self.font_label).pack(side=tk.LEFT, padx=(0, 8))

        self.date_filter = ttk.Combobox(filter_frame, width=22, state='readonly', font=self.font_entry)
        self.date_filter.pack(side=tk.LEFT, padx=(0, 15))
        self.date_filter.bind('<<ComboboxSelected>>', lambda e: self.filter_completed())

        # Toolbar
        toolbar = tk.Frame(parent)
        toolbar.pack(fill=tk.X, pady=(0, 10))

        tk.Button(toolbar, text=self.T['btn_del_completed'], command=self.delete_completed,
                  bg="#E67E22", fg="white", font=self.font_button,
                  cursor="hand2", padx=16, pady=4).pack(side=tk.LEFT, padx=(0, 8))

        tk.Button(toolbar, text=self.T['btn_clear'], command=self.delete_all_completed,
                  bg="#C0392B", fg="white", font=self.font_button,
                  cursor="hand2", padx=16, pady=4).pack(side=tk.LEFT, padx=(0, 16))

        tk.Button(toolbar, text=self.T['btn_export'], command=self.export_completed,
                  bg=THEME['colors']['primary'], fg="white", font=self.font_button,
                  cursor="hand2", padx=16, pady=4).pack(side=tk.LEFT, padx=(0, 8))

        
        # ── Sorting Dropdown ──
        sort_frame = tk.Frame(toolbar)
        sort_frame.pack(side=tk.LEFT, padx=(20, 0))
        
        tk.Label(sort_frame, text="排序：", font=self.font_label).pack(side=tk.LEFT, padx=(0, 5))
        self.sort_combobox_completed = ttk.Combobox(sort_frame, textvariable=self.sort_criteria, 
                                                   values=[self.T['sort_default'], self.T['sort_priority'], self.T['sort_created'], self.T['sort_desc'], self.T['sort_completed']],
                                                   state="readonly", width=12, font=self.font_label)
        self.sort_combobox_completed.pack(side=tk.LEFT)
        self.sort_combobox_completed.bind('<<ComboboxSelected>>', lambda e: self.refresh_completed())
        
        # ── Search Box ──
        search_frame = tk.Frame(toolbar)
        search_frame.pack(side=tk.LEFT, padx=(20, 0))
        
        tk.Label(search_frame, text="🔍 搜索：", font=self.font_label).pack(side=tk.LEFT, padx=(0, 5))
        self.search_entry_completed = ttk.Entry(search_frame, textvariable=self.search_query, width=15, font=self.font_label)
        self.search_entry_completed.pack(side=tk.LEFT)
        
        # Table
        table_frame = tk.Frame(parent)
        table_frame.pack(fill=tk.BOTH, expand=True)

        columns = ('☑️', '编号', self.T['col_desc'], self.T['col_type'], self.T['col_priority'], self.T['col_owner'], self.T['col_tag'], self.T['col_progress'], self.T['col_complete_time'], self.T['col_duration'])
        self.completed_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=20)
        
        # 已完成任务背景色配置
        # 当天完成:深灰色 | 超过1天:浅灰色
        self.completed_tree.tag_configure('completed_today', background='#A9A9A9', foreground='#666666')
        self.completed_tree.tag_configure('completed_old', background='#D3D3D3', foreground='#666666')
        
        col_widths = {
            '☑️': 35, '编号': 40, self.T['col_desc']: 260,
            self.T['col_type']: 100, self.T['col_priority']: 100, self.T['col_owner']: 100,
            self.T['col_tag']: 90, self.T['col_progress']: 70,
            self.T['col_complete_time']: 130, self.T['col_duration']: 100
        }
        col_minwidths = {
            '☑️': 30, '编号': 35, self.T['col_desc']: 260,
            self.T['col_type']: 80, self.T['col_priority']: 80, self.T['col_owner']: 80,
            self.T['col_tag']: 70, self.T['col_progress']: 60,
            self.T['col_complete_time']: 110, self.T['col_duration']: 90
        }
        for col in columns:
            self.completed_tree.heading(col, text=col, anchor=tk.CENTER)
            # 方案B（V1.6.21 同步）：仅「冒险描述」stretch=True 吸收伸缩，其余列全部钉死
            # → 全屏后缩小 = 首开小窗口效果
            anchor = tk.W if col in [self.T['col_owner'], self.T['col_tag']] else tk.CENTER
            w = col_widths.get(col, 100)
            mw = col_minwidths.get(col, 40)
            if col == self.T['col_desc']:
                self.completed_tree.column(col, width=w, minwidth=mw, stretch=True, anchor=anchor)
            else:
                self.completed_tree.column(col, stretch=False, width=w, minwidth=mw, anchor=anchor)

        vsb = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.completed_tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.completed_tree.xview)
        self.completed_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.completed_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
        
        # 绑定右键菜单
        self.completed_tree.bind('<Button-3>', self.show_completed_context_menu)
        self.completed_context_menu = tk.Menu(self.root, tearoff=0, font=self.font_label)
        self.completed_context_menu.add_command(label=self.T['ctx_restore'], command=self.restore_task)
        self.completed_context_menu.add_command(label=self.T['ctx_del_completed'], command=self.delete_completed)
        self.completed_context_menu.add_command(label=self.T['btn_export'], command=self.export_completed)
        self.completed_context_menu.add_command(label=self.T['btn_clear'], command=self.delete_all_completed)
    
    def show_completed_context_menu(self, event):
        """Show right-click context menu for completed tasks"""
        try:
            # Select the item under cursor
            item = self.completed_tree.identify_row(event.y)
            if item:
                self.completed_tree.selection_set(item)
                self.completed_context_menu.post(event.x_root, event.y_root)
        except Exception as e:
            print(f"Error showing completed context menu: {e}")
    
    def setup_stats_tab(self, parent):
        """Setup statistics tab with growth system panels + overview tables"""
        # ── 可滚动容器：解决战绩内容超长无法滚轮查看 ──
        stats_canvas = tk.Canvas(parent, bg=THEME['colors']['panel_bg'], highlightthickness=0)
        stats_scrollbar = ttk.Scrollbar(parent, orient="vertical",
                                        command=stats_canvas.yview)
        stats_canvas.configure(yscrollcommand=stats_scrollbar.set)
        stats_scrollbar.pack(side="right", fill="y")
        stats_canvas.pack(side="left", fill="both", expand=True)

        container = tk.Frame(stats_canvas, bg=THEME['colors']['panel_bg'])
        container_id = stats_canvas.create_window((0, 0), window=container, anchor="nw")

        def _sync_stats_width(event=None):
            stats_canvas.itemconfig(container_id, width=stats_canvas.winfo_width())
            stats_canvas.configure(scrollregion=stats_canvas.bbox("all"))

        container.bind("<Configure>", lambda e: stats_canvas.configure(
            scrollregion=stats_canvas.bbox("all")))
        stats_canvas.bind("<Configure>", _sync_stats_width)

        def _on_stats_mousewheel(event):
            # 仅当「冒险战绩」tab 可见时才滚动，避免影响其他 tab
            try:
                if str(self.notebook.select()) == str(parent):
                    stats_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            except Exception:
                pass

        stats_canvas.bind_all("<MouseWheel>", _on_stats_mousewheel)
        # ── 0. 成长系统面板（顶部，两个独立面板）──
        growth_outer = tk.Frame(container, bg=THEME['colors']['panel_bg'])
        growth_outer.pack(fill=tk.X, padx=10, pady=(10, 8))
        
        growth_frame = tk.Frame(growth_outer, bg=THEME['colors']['panel_bg'])
        growth_frame.pack(fill=tk.X)
        
        # ── 0a. 本周黄金之路面板（V1.6.0：移除金币系统，改为每周智慧）──
        exp_card = tk.Frame(growth_frame, bg=THEME['colors']['title_bar'])
        exp_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        tk.Label(exp_card, text=self.T['exp_card_title'], font=("Microsoft YaHei", 13, "bold"),
                 bg=THEME['colors']['title_bar'], fg=THEME['colors']['status_bar']).pack(anchor=tk.W)
        
        self.exp_rank_label = tk.Label(
            exp_card, text=f"{self.get_wisdom_title(1)} | Lv.1", font=("Microsoft YaHei", 22, "bold"),
            bg=THEME['colors']['title_bar'], fg=THEME['colors']['accent']
        )
        self.exp_rank_label.pack(pady=(6, 2))
        
        # 本周进度（剩余天数）
        self.exp_daily_label = tk.Label(
            exp_card, text=self.T['exp_daily'].format(1, 6), font=("Microsoft YaHei", 12, "bold"),
            bg=THEME['colors']['title_bar'], fg="#2ECC71"
        )
        self.exp_daily_label.pack(pady=(0, 4))
        
        # 智慧进度条（Canvas绘制）
        self.exp_progress_canvas = tk.Canvas(exp_card, width=320, height=24,
                                              bg=THEME['colors']['dark_alt'], highlightthickness=0)
        self.exp_progress_canvas.pack(pady=(4, 2))
        
        self.exp_detail_label = tk.Label(
            exp_card, text=f"0/{self.get_wisdom_for_level(2)[1]} 智慧 → {self.get_wisdom_title(2)} (Lv.2)", font=("Microsoft YaHei", 10),
            bg=THEME['colors']['title_bar'], fg="#BDC3C7"
        )
        self.exp_detail_label.pack(anchor=tk.W, pady=(2, 0))
        
        # ── 1. Week filter ──
        filter_frame = tk.Frame(container)
        filter_frame.pack(fill=tk.X, padx=10, pady=(10, 5))
        
        tk.Label(filter_frame, text="选择周:", font=self.font_label).pack(side=tk.LEFT, padx=(0, 8))
        
        self.week_filter_var = tk.StringVar(value="")
        self.week_filter_combo = ttk.Combobox(filter_frame, textvariable=self.week_filter_var,
                                               width=22, state='readonly', font=self.font_entry)
        self.week_filter_combo.pack(side=tk.LEFT, padx=(0, 15))
        self.week_filter_combo.bind('<<ComboboxSelected>>',
                                    lambda e: self.refresh_stats())
        
        # ── 实时时钟（选择周 右侧显示，描边幽灵牌：浅灰底+蓝描边+蓝色线条时钟+蓝字）──
        self.clock_frame = tk.Frame(filter_frame, bg=THEME['colors']['panel_bg'],
                                    highlightbackground=THEME['colors']['primary'], highlightthickness=1,
                                    bd=0, relief="flat")
        self.clock_frame.pack(side=tk.RIGHT, padx=(10, 0))
        # 蓝色线条时钟图标（Canvas 绘制，规避 emoji 单色限制，颜色可控与描边同色）
        self.clock_canvas = tk.Canvas(self.clock_frame, width=16, height=16,
                                      bg=THEME['colors']['panel_bg'], highlightthickness=0)
        self.clock_canvas.pack(side=tk.LEFT, padx=(6, 2))
        # 时间文字（深蓝，与界面主色一致）
        self.clock_text = tk.Label(self.clock_frame, text="", fg=THEME['colors']['title_bar'],
                                   bg=THEME['colors']['panel_bg'], font=self.font_label)
        self.clock_text.pack(side=tk.LEFT, padx=(0, 6))
        self._update_clock()
        
        # ── 1. Overview cards ──
        overview_frame = tk.LabelFrame(container, text="📊 总览",
                                       font=self.font_label, padx=15, pady=12)
        overview_frame.pack(fill=tk.X, padx=10, pady=(0, 8))
        
        card_frame = tk.Frame(overview_frame)
        card_frame.pack(fill=tk.X)
        
        self.stat_pending = tk.StringVar(value="0")
        self.stat_completed = tk.StringVar(value="0")
        self.stat_rate = tk.StringVar(value="0%")
        
        card_data = [
            (self.T['overview_pending'], self.stat_pending, "#F0AD4E"),
            (self.T['overview_done'], self.stat_completed, "#5CB85C"),
            (self.T['overview_rate'], self.stat_rate, THEME['colors']['primary']),
        ]
        for i, (label, var, bg) in enumerate(card_data):
            card = tk.Frame(card_frame, bg=bg, padx=20, pady=8)
            card.pack(side=tk.LEFT, fill=tk.X, expand=True,
                      padx=(0, 10) if i < len(card_data) - 1 else 0)
            tk.Label(card, text=label, font=self.font_label,
                     bg=bg, fg="white").pack()
            tk.Label(card, textvariable=var,
                     font=("Microsoft YaHei", 28, "bold"),
                     bg=bg, fg="white").pack()
        
        # ── 2. By Priority + By Category side by side ──
        mid_frame = tk.Frame(container)
        mid_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
        
        # -- Priority table --
        pri_frame = tk.LabelFrame(mid_frame, text=self.T['stat_priority'],
                                  font=self.font_label, padx=8, pady=8)
        pri_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        
        p_cols = (self.T['col_priority'], self.T['stat_ongoing'], self.T['stat_done'], self.T['stat_total'])
        pri_widths = {self.T['col_priority']: 105, self.T['stat_ongoing']: 95,
                       self.T['stat_done']: 95, self.T['stat_total']: 85}
        self.stat_priority_tree = ttk.Treeview(pri_frame, columns=p_cols,
                                               show='headings', height=5)
        for c in p_cols:
            self.stat_priority_tree.heading(c, text=c, anchor=tk.CENTER)
            self.stat_priority_tree.column(c, width=pri_widths.get(c, 95), minwidth=70, anchor=tk.CENTER)
        self.stat_priority_tree.pack(fill=tk.BOTH, expand=True)
        
        # -- Category table --
        cat_frame = tk.LabelFrame(mid_frame, text=self.T['stat_category'],
                                  font=self.font_label, padx=8, pady=8)
        cat_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        
        c_cols = (self.T['col_type'], self.T['stat_ongoing'], self.T['stat_done'], self.T['stat_total'])
        self.stat_category_tree = ttk.Treeview(cat_frame, columns=c_cols,
                                               show='headings', height=5)
        col_widths = {self.T['col_type']: 135, self.T['stat_ongoing']: 95,
                       self.T['stat_done']: 95, self.T['stat_total']: 85}
        for c in c_cols:
            self.stat_category_tree.heading(c, text=c, anchor=tk.CENTER)
            self.stat_category_tree.column(c, width=col_widths.get(c, 95), minwidth=80, anchor=tk.CENTER)
        self.stat_category_tree.pack(fill=tk.BOTH, expand=True)
        
        # ── 3. By Date ──
        date_frame = tk.LabelFrame(container, text=self.T['date_stat_title'],
                                   font=self.font_label, padx=8, pady=8)
        date_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))

        # 日期筛选下拉（默认全部，可选具体某天）
        dfilter_row = tk.Frame(date_frame)
        dfilter_row.pack(fill=tk.X, pady=(0, 6))
        tk.Label(dfilter_row, text="筛选日期：", font=self.font_entry).pack(side=tk.LEFT)
        self.stat_date_filter = ttk.Combobox(dfilter_row, width=20, state='readonly',
                                             font=self.font_entry)
        self.stat_date_filter.pack(side=tk.LEFT, padx=(6, 0))
        self.stat_date_filter.bind('<<ComboboxSelected>>',
                                   lambda e: self._refresh_stat_date())

        d_cols = (self.T['date_col'], self.T['done_count_col'])
        self.stat_date_tree = ttk.Treeview(date_frame, columns=d_cols,
                                           show='headings')
        for c in d_cols:
            self.stat_date_tree.heading(c, text=c, anchor=tk.CENTER)
            self.stat_date_tree.column(c, width=160, minwidth=100, anchor=tk.CENTER)
        self.stat_date_tree.pack(fill=tk.BOTH, expand=True)
    

    def add_task(self):
        """Add a new task"""
        try:
            task_text = self.task_entry.get().strip()
            
            # Validate task text
            is_valid, error_msg = self.validate_task_text(task_text)
            if not is_valid:
                messagebox.showwarning("警告", error_msg)
                return
            
            priority = self.priority_var.get()
            category = self.category_var.get()
            person = self.person_entry.get().strip()
            
            # Validate person name
            is_valid, error_msg = self.validate_person_name(person)
            if not is_valid:
                messagebox.showwarning("警告", error_msg)
                return
            
            if not person:
                person = "韦程"
            
            # Process and validate tags
            tags_text = self.tags_entry.get().strip()
            is_valid, error_msg, tags = self.validate_tags(tags_text)
            if not is_valid:
                messagebox.showwarning("警告", error_msg)
                return
            
            # Calculate new unique ID
            all_tasks = self.data["tasks"] + self.data["completed"]
            all_ids = [t["id"] for t in all_tasks if "id" in t]
            new_id = max(all_ids) + 1 if all_ids else 1
            
            task = {
                "id": new_id,
                "text": task_text,
                "priority": priority,
                "category": category,
                "person": person,
                "tags": tags,  # Add tags field
                "created": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            
            self.data["tasks"].append(task)
            self.save_data()
            self.task_entry.delete(0, tk.END)
            self.person_entry.delete(0, tk.END)
            self.tags_entry.delete(0, tk.END)  # Clear tags input
            self.refresh_display()
            
            # Update game data
            self.game_data["stats"]["total_added"] += 1
            self.save_game_data()
            self.update_game_display()
            
            # Show new task animation
            self.show_new_task_animation()
            
            self.set_status(self.T['status_task_received'].format(task_text))
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(error_detail)
            messagebox.showerror(self.T['err_new_task'], f"错误详情:\n{str(e)}")
    
    def _complete_task_by_id(self, task_id):
        """标记指定任务为完成：移动数据 + 更新游戏数据 + 播放动画。
        供工具栏「标记通关」与表格复选框点击共用，消除重复逻辑。"""
        task = None
        for i, t in enumerate(self.data["tasks"]):
            if t["id"] == task_id:
                task = t
                t["completed"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                self.data["completed"].append(t)
                self.data["tasks"].pop(i)
                break
        if not task:
            return False

        # Update game data
        self.game_data["stats"]["total_completed"] += 1
        priority = task.get("priority", "")
        amount = self.WISDOM_BY_PRIORITY.get(priority, 5)  # 默认5智慧
        task["wisdom_gain"] = amount  # 记录贡献分，随任务进入 completed 保留
        self.add_wisdom(amount)
        self.save_game_data()
        self.update_game_display()

        # 动画（与完成语义绑定，统一在此播放）
        self.show_wisdom_animation(amount)
        self.show_fireworks_animation()
        return True

    def complete_task(self):
        """Mark a task as completed"""
        try:
            selection = self.pending_tree.selection()
            if not selection:
                messagebox.showwarning("警告", self.T['warn_select_complete'])
                return
            
            task_id = int(selection[0])
            if self._complete_task_by_id(task_id):
                self.save_data()
                self.refresh_display()
                self.set_status(self.T['status_task_done'])
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            messagebox.showerror("操作出错", f"错误详情:\n{str(e)}")
    
    def on_checkbox_click(self, event):
        """Handle checkbox click in pending tasks - mark task as complete or toggle subtask"""
        try:
            # Get the clicked item and column
            item = self.pending_tree.identify_row(event.y)
            column = self.pending_tree.identify_column(event.x)
            
            # #0 树列（折叠三角 ▸/▾）仅负责展开/折叠，不进入完成逻辑
            if column == '#0':
                return
            
            # 复选框位于第一个数据列 #1（#0 为树形折叠三角列，固定不后移）
            if column == '#1' and item:
                # Check if this is a subtask (iid contains "_subtask_")
                if "_subtask_" in item:
                    # Parse subtask iid: "{task_id}_subtask_{subtask_idx}"
                    parts = item.split("_subtask_")
                    task_id = int(parts[0])
                    subtask_idx = int(parts[1]) - 1  # Convert to 0-based index
                    
                    # Find the task and toggle subtask
                    for task in self.data["tasks"]:
                        if task["id"] == task_id:
                            if "subtasks" in task and subtask_idx < len(task["subtasks"]):
                                # Toggle subtask done status
                                task["subtasks"][subtask_idx]["done"] = not task["subtasks"][subtask_idx].get("done", False)
                                self.save_data()
                                self.refresh_display()
                                self.set_status("✓ 支线任务状态已更新！")
                            break
                else:
                    # Main task - mark as complete
                    task_id = int(item)
                    if self._complete_task_by_id(task_id):
                        self.save_data()
                        self.refresh_display()
                        self.set_status(self.T['status_task_done'])
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            messagebox.showerror("操作出错", f"错误详情:\n{str(e)}")
    
    def delete_task(self):
        """Delete a pending task with confirmation"""
        selection = self.pending_tree.selection()
        if not selection:
            messagebox.showwarning("警告", self.T['warn_select_abandon'])
            return
        
        task_id = int(selection[0])
        # Find task text for confirmation
        task_text = ""
        for t in self.data["tasks"]:
            if t["id"] == task_id:
                task_text = t.get("text", "")
                break
        
        if not messagebox.askyesno("确认放弃", self.T['confirm_abandon'].format(task_text)):
            return
        
        self.data["tasks"] = [t for t in self.data["tasks"] if t["id"] != task_id]
        self.save_data()
        self.refresh_display()
        self.set_status(self.T['status_abandoned'])
    
    def edit_task(self):
        """Edit selected pending task's content, category, priority, or person"""
        selection = self.pending_tree.selection()
        if not selection:
            messagebox.showwarning("警告", self.T['warn_select_edit'])
            return
        
        task_id = int(selection[0])
        task = None
        for t in self.data["tasks"]:
            if t["id"] == task_id:
                task = t
                break
        
        if not task:
            messagebox.showerror("错误", self.T['err_not_found_edit'])
            return
        
        # ── Build edit dialog ──
        dialog = tk.Toplevel(self.root)
        dialog.title(f"{self.T['edit_dlg_title']}{task_id}")
        dialog.geometry("520x450")  # Increased height to show all fields
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center the dialog
        dialog.update_idletasks()
        width = dialog.winfo_width()
        height = dialog.winfo_height()
        x = (dialog.winfo_screenwidth() // 2) - (width // 2)
        y = (dialog.winfo_screenheight() // 2) - (height // 2)
        dialog.geometry(f"{width}x{height}+{x}+{y}")
        
        main_frame = tk.Frame(dialog, padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # --- 📝 冒险描述 ---
        tk.Label(main_frame, text=self.T['lbl_desc_edit'], font=self.font_label).pack(anchor=tk.W, pady=(0, 4))
        text_var = tk.StringVar(value=task.get("text", ""))
        text_entry = tk.Entry(main_frame, font=self.font_entry, textvariable=text_var)
        text_entry.pack(fill=tk.X, pady=(0, 14))
        text_entry.focus_set()
        
        # --- 🗺️ 冒险类型 ---
        row_frame = tk.Frame(main_frame)
        row_frame.pack(fill=tk.X, pady=(0, 14))
        
        g1 = tk.Frame(row_frame)
        g1.pack(side=tk.LEFT, padx=(0, 20))
        tk.Label(g1, text=self.T['lbl_type'], font=self.font_label).pack(anchor=tk.W, pady=(0, 4))
        category_var = tk.StringVar(value=task.get("category", self.category_options[0]))
        category_combo = ttk.Combobox(g1, textvariable=category_var,
                                      values=self.category_options,
                                      state='readonly', width=16, font=self.font_entry)
        category_combo.pack()
        
        g2 = tk.Frame(row_frame)
        g2.pack(side=tk.LEFT)
        tk.Label(g2, text=self.T['lbl_priority'], font=self.font_label).pack(anchor=tk.W, pady=(0, 4))
        priority_var = tk.StringVar(value=task.get("priority", self.priority_options[0]))
        priority_combo = ttk.Combobox(g2, textvariable=priority_var,
                                      values=self.priority_options,
                                      state='readonly', width=16, font=self.font_entry)
        priority_combo.pack()
        
        # --- ⚔️ 冒险者 ---
        tk.Label(main_frame, text=self.T['lbl_owner'], font=self.font_label).pack(anchor=tk.W, pady=(0, 4))
        person_var = tk.StringVar(value=task.get("person", ""))
        person_entry = tk.Entry(main_frame, font=self.font_entry, textvariable=person_var)
        person_entry.pack(fill=tk.X, pady=(0, 14))
        
        # --- 🏷️ 标签 ---
        tk.Label(main_frame, text=self.T['dlg_tag2'], font=self.font_label).pack(anchor=tk.W, pady=(0, 4))
        tags_var = tk.StringVar(value=", ".join(task.get("tags", [])))
        tags_entry = tk.Entry(main_frame, font=self.font_entry, textvariable=tags_var)
        tags_entry.pack(fill=tk.X, pady=(0, 18))
        
        # --- Buttons ---
        btn_frame = tk.Frame(main_frame)
        btn_frame.pack(fill=tk.X)
        
        def save_edit():
            """Save the edited task."""
            new_text = text_var.get().strip()
            
            # Validate task text
            is_valid, error_msg = self.validate_task_text(new_text)
            if not is_valid:
                messagebox.showwarning("警告", error_msg)
                return
            
            task["text"] = new_text
            task["category"] = category_var.get()
            task["priority"] = priority_var.get()
            person_val = person_var.get().strip()
            
            # Validate person name
            is_valid, error_msg = self.validate_person_name(person_val)
            if not is_valid:
                messagebox.showwarning("警告", error_msg)
                return
            
            task["person"] = person_val if person_val else "韦程"
            
            # Validate tags
            tags_text = tags_var.get().strip()
            is_valid, error_msg, tags = self.validate_tags(tags_text)
            if not is_valid:
                messagebox.showwarning("警告", error_msg)
                return
            
            task["tags"] = tags
            
            self.save_data()
            self.refresh_display()
            self.set_status(self.T['status_task_updated'].format(new_text))
            dialog.destroy()
        
        tk.Button(btn_frame, text="✚ 保存修改", command=save_edit,
                  bg=THEME['colors']['primary'], fg="white", font=self.font_button,
                  cursor="hand2", padx=20, pady=4).pack(side=tk.RIGHT, padx=(10, 0))
        
        tk.Button(btn_frame, text="取消", command=dialog.destroy,
                  font=self.font_button,
                  cursor="hand2", padx=20, pady=4).pack(side=tk.RIGHT)
        self._apply_emoji_images(dialog)
    
    def restore_task(self):
        """Restore a completed task back to pending"""
        try:
            selection = self.completed_tree.selection()
            if not selection:
                messagebox.showwarning("警告", self.T['warn_select_restore'])
                return
            
            task_id = int(selection[0])
            
            # Find the task in completed list
            task_to_restore = None
            for i, task in enumerate(self.data["completed"]):
                if task["id"] == task_id:
                    task_to_restore = task
                    break
            
            if not task_to_restore:
                messagebox.showerror("错误", "未找到该任务！")
                return
            
            # 计算需扣回的智慧（恢复=撤销完成）
            restore_wisdom = task_to_restore.get("wisdom_gain", 0)
            is_this_week = self.is_task_this_week(
                task_to_restore.get("completed", ""),
                self.game_data.get("week_start_date", "")
            )
            
            # Remove completed timestamp
            if "completed" in task_to_restore:
                del task_to_restore["completed"]
            
            # Move task back to pending
            self.data["tasks"].append(task_to_restore)
            self.data["completed"].remove(task_to_restore)
            
            self.save_data()
            self.refresh_display()
            # 同步扣回智慧
            if restore_wisdom > 0:
                self.subtract_wisdom(restore_wisdom, restore_wisdom if is_this_week else 0)
            # V1.6.2：恢复=撤销完成，total_completed 同步回退，保持统计一致
            self.game_data["stats"]["total_completed"] = max(
                0, self.game_data["stats"]["total_completed"] - 1)
            self.save_game_data()
            self.set_status(self.T['status_task_restored'].format(task_to_restore.get('text', ''), restore_wisdom))
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            messagebox.showerror("操作出错", f"错误详情:\n{str(e)}")
    
    def delete_completed(self):
        """Delete a completed task record with confirmation"""
        selection = self.completed_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "请选择要删除的记录！")
            return
        
        task_id = int(selection[0])
        # Find task info for confirmation
        task_text = ""
        task_wisdom = 0
        task_completed = ""
        for t in self.data["completed"]:
            if t["id"] == task_id:
                task_text = t.get("text", "")
                task_wisdom = t.get("wisdom_gain", 0)
                task_completed = t.get("completed", "")
                break
        
        is_this_week = self.is_task_this_week(task_completed, self.game_data.get("week_start_date", ""))
        if not messagebox.askyesno("确认删除",
            self.T['confirm_del_done'].format(task_text, task_wisdom)):
            return
        
        self.data["completed"] = [t for t in self.data["completed"] if t["id"] != task_id]
        self.save_data()
        self.refresh_completed()
        # 同步扣回智慧（本周任务同时扣 weekly，历史任务只扣 total）
        self.subtract_wisdom(task_wisdom, task_wisdom if is_this_week else 0)
        self.set_status(f"🗑️ 记录已删除，已扣回 {task_wisdom} 智慧")
    
    def delete_all_completed(self):
        """Delete all completed task records with confirmation"""
        count = len(self.data["completed"])
        if count == 0:
            self.set_status("🏁 已通关列表为空，无需删除")
            return
        
        # 汇总扣回（本周任务计入 weekly，全部计入 total）
        total_d = 0
        weekly_d = 0
        week_start = self.game_data.get("week_start_date", "")
        for t in self.data["completed"]:
            w = t.get("wisdom_gain", 0)
            total_d += w
            if self.is_task_this_week(t.get("completed", ""), week_start):
                weekly_d += w
        
        if not messagebox.askyesno("确认删除全部", 
            f"🗑️ 确定要删除全部 {count} 条已通关记录吗？\n\n"
            f"⚠️ 此操作将扣回共 {total_d} 智慧（其中本周 {weekly_d}），本周等级可能下降\n此操作不可撤销！"):
            return
        
        self.data["completed"] = []
        self.save_data()
        self.refresh_completed()
        self.refresh_display()
        # 同步扣回
        self.subtract_wisdom(total_d, weekly_d)
        self.set_status(f"🗑️ 已删除全部，共扣回 {total_d} 智慧")
    
    def export_completed(self):
        """Export all completed tasks to Excel file"""
        if not self.data["completed"]:
            self.set_status("已完成列表为空，无可导出的记录")
            return
        
        now = datetime.now()
        iso_year, iso_week, _ = now.isocalendar()
        
        # If week filter is active, include week info in filename
        selected_week = self.week_filter_var.get()
        if selected_week and selected_week != "全部":
            default_filename = self.T['export_file_week'].format(week=selected_week)
        else:
            default_filename = self.T['export_file_iso'].format(year=iso_year, week=iso_week)
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
            title=self.T['btn_export'],
            initialfile=default_filename
        )
        if not file_path:
            return  # user cancelled
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.T['export_sheet']
        
        # ── Header style ──
        header_font = Font(name="Microsoft YaHei", bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        
        # ── Write headers ──
        headers = [self.T['col_id_emoji'], self.T['col_desc'], self.T['col_type'], self.T['col_priority'], self.T['col_owner'], self.T['col_tag'], self.T['col_complete_time'], self.T['col_receive'], self.T['col_duration']]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        
        # ── Write data ──
        data_font = Font(name="Microsoft YaHei", size=11)
        data_align = Alignment(horizontal="center", vertical="center")
        content_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        for row_idx, task in enumerate(self.data["completed"], 2):
            duration = self.calculate_duration(task.get("created", ""), task.get("completed", ""))
            tags_str = ", ".join(task.get("tags", [])) if task.get("tags") else ""
            values = [
                row_idx - 1,
                task.get("text", ""),
                task.get("category", ""),
                task.get("priority", ""),
                self.format_person(task.get("person", "")),
                tags_str,  # Add tags
                task.get("completed", ""),
                task.get("created", ""),
                duration
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.border = thin_border
                if col_idx == 2:  # 描述列左对齐
                    cell.alignment = content_align
                else:
                    cell.alignment = data_align
        
        # ── Auto-fit column widths ──
        col_widths_map = {1: 6, 2: 45, 3: 14, 4: 14, 5: 16, 6: 15, 7: 18, 8: 18, 9: 15}
        for col_idx, width in col_widths_map.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
        
        # Freeze top row
        ws.freeze_panes = "A2"
        
        try:
            wb.save(file_path)
            self.set_status(f"✓ 导出成功: {os.path.basename(file_path)} ({len(self.data['completed'])} 条记录)")
        except Exception as e:
            messagebox.showerror("导出失败", f"保存文件失败:\n{str(e)}")
    
    # ── Timer Methods (Compact Timer) ──
    def reset_timer(self):
        """Reset timer to default time and stop."""
        # Stop current timer if running
        self.timer_running = False
        self.timer_paused = False
        
        if self.timer_after_id:
            self.root.after_cancel(self.timer_after_id)
            self.timer_after_id = None
        
        # Reset to default time
        self.timer_seconds = self.timer_default_seconds
        self.timer_display_var.set(self.format_timer_display())
        
        # Reset button
        self.timer_btn.config(text="▶ 开始", bg="#3498DB")
    
    def toggle_timer(self):
        """Start or pause the timer."""
        if self.timer_running and not self.timer_paused:
            # Pause timer
            self.timer_paused = True
            self.timer_btn.config(text="▶ 继续", bg="#F39C12")
            if self.timer_after_id:
                self.root.after_cancel(self.timer_after_id)
                self.timer_after_id = None
        elif self.timer_running and self.timer_paused:
            # Resume timer
            self.timer_paused = False
            self.timer_btn.config(text="⏸ 暂停", bg="#E74C3C")
            self.update_timer()
        else:
            # Start timer
            self.timer_running = True
            self.timer_paused = False
            self.timer_btn.config(text="⏸ 暂停", bg="#E74C3C")
            self.update_timer()
    
    def update_timer(self):
        """Update timer display and check if time's up."""
        if self.timer_seconds <= 0:
            # Time's up
            self.timer_running = False
            self.timer_paused = False
            self.timer_btn.config(text="▶ 开始", bg="#3498DB")
            self.timer_display_var.set("00:00")
            
            # Show different message based on timer duration
            if self.timer_default_seconds <= 15 * 60:  # 5 or 15 minutes
                messagebox.showinfo("⏰ 时间到！", "⏰ 番茄钟结束！该搬砖了~")
            else:  # 25 or 30 minutes
                messagebox.showinfo("⏰ 时间到！", "⏰ 番茄钟结束！休息一下吧~")
            
            # Reset to default time
            self.timer_seconds = self.timer_default_seconds
            self.timer_display_var.set(self.format_timer_display())
            return
        
        # Update display
        self.timer_display_var.set(self.format_timer_display())
        
        # Decrease seconds
        self.timer_seconds -= 1
        
        # Schedule next update
        self.timer_after_id = self.root.after(1000, self.update_timer)
    
    def format_timer_display(self):
        """Format seconds to MM:SS display."""
        minutes = self.timer_seconds // 60
        seconds = self.timer_seconds % 60
        return f"⏰ {minutes:02d}:{seconds:02d}"
    
    def set_timer(self, minutes):
        """Set timer to specified minutes."""
        # Stop current timer if running
        if self.timer_running:
            if self.timer_after_id:
                self.root.after_cancel(self.timer_after_id)
                self.timer_after_id = None
            self.timer_running = False
            self.timer_paused = False
            self.timer_btn.config(text="▶ 开始", bg="#3498DB")
        
        # Set new time
        self.timer_seconds = minutes * 60
        self.timer_default_seconds = minutes * 60
        self.timer_display_var.set(self.format_timer_display())
    
    def show_timer_menu(self, event):
        """Show timer right-click menu to change time."""
        self.timer_menu.post(event.x_root, event.y_root)
    
    def reload_data(self):
        """Reload data from JSON file and refresh display."""
        try:
            self.data = self.load_data()
            self.refresh_display()
            self.set_status("✓ 数据已刷新")
        except Exception as e:
            messagebox.showerror("刷新失败", f"重新加载数据失败:\n{str(e)}")
    
    def show_add_subtask_dialog(self, task):
        """Show dialog to add a new subtask directly."""
        # Create dialog
        dialog = tk.Toplevel(self.root)
        dialog.title(self.T['dlg_title_add'])  # 原"➕ 添加支线任务"
        dialog.geometry("400x210")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        # Don't use grab_set() to avoid focus issues
        
        # Center the dialog
        dialog.update_idletasks()
        w = dialog.winfo_width()
        h = dialog.winfo_height()
        x = (dialog.winfo_screenwidth() // 2) - (w // 2)
        y = (dialog.winfo_screenheight() // 2) - (h // 2)
        dialog.geometry(f"{w}x{h}+{x}+{y}")
        
        f = tk.Frame(dialog, padx=20, pady=20)
        f.pack(fill=tk.BOTH, expand=True)
        
        tk.Label(f, text="支线任务描述:", font=self.font_label).pack(anchor=tk.W, pady=(0, 8))

        text_var = tk.StringVar()
        entry = tk.Entry(f, font=self.font_entry, textvariable=text_var)
        entry.pack(fill=tk.X, pady=(0, 12))
        
        # 标签（逗号分隔）
        tk.Label(f, text=self.T['dlg_tag'], font=self.font_label).pack(anchor=tk.W, pady=(0, 4))
        tags_var = tk.StringVar()
        tags_entry = tk.Entry(f, font=self.font_entry, textvariable=tags_var)
        tags_entry.pack(fill=tk.X, pady=(0, 12))
        # Use after() to ensure focus is set after dialog is fully created
        entry.after(100, lambda: entry.focus_force())
        
        def save_subtask():
            """Save the new subtask."""
            text = text_var.get().strip()
            if not text:
                messagebox.showwarning("警告", "支线任务描述不能为空！")
                return
            
            # 校验标签
            tags_text = tags_var.get().strip()
            is_valid, error_msg, tags = self.validate_tags(tags_text)
            if not is_valid:
                messagebox.showwarning("警告", error_msg)
                return
            
            # Add subtask to task
            if "subtasks" not in task:
                task["subtasks"] = []
            
            new_subtask = {"id": len(task["subtasks"]) + 1, "text": text, "done": False,
                          "tags": tags, "created": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
            task["subtasks"].append(new_subtask)
            
            # Save data and refresh display
            self.save_data()
            self.refresh_display()
            
            # Close dialog
            dialog.destroy()
            
            self.set_status("✓ 支线任务已添加！")
        
        # Buttons
        bf = tk.Frame(f)
        bf.pack(fill=tk.X)
        
        tk.Button(bf, text="✅ 保存", command=save_subtask,
                  bg="#5CB85C", fg="white", font=self.font_button,
                  cursor="hand2", padx=16, pady=4).pack(side=tk.RIGHT, padx=(8, 0))
        
        tk.Button(bf, text="取消", command=dialog.destroy,
                  font=self.font_button,
                  cursor="hand2", padx=16, pady=4).pack(side=tk.RIGHT)
    
    def manage_subtasks(self, task):
        """Manage subtasks for a task."""
        # ── Build subtasks management dialog ──
        dialog = tk.Toplevel(self.root)
        dialog.title(f"📋 管理支线任务 - {task['text'][:20]}...")
        dialog.geometry("500x400")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center the dialog
        dialog.update_idletasks()
        width = dialog.winfo_width()
        height = dialog.winfo_height()
        x = (dialog.winfo_screenwidth() // 2) - (width // 2)
        y = (dialog.winfo_screenheight() // 2) - (height // 2)
        dialog.geometry(f"{width}x{height}+{x}+{y}")
        
        main_frame = tk.Frame(dialog, padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # --- Subtasks list ---
        tk.Label(main_frame, text="📋 支线任务列表:", font=self.font_label).pack(anchor=tk.W, pady=(0, 8))
        
        # Listbox with scrollbar
        list_frame = tk.Frame(main_frame)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 12))
        
        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        listbox = tk.Listbox(list_frame, font=self.font_entry, height=10,
                              selectmode=tk.SINGLE, yscrollcommand=scrollbar.set)
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=listbox.yview)
        
        # Populate listbox
        subtasks = task.get("subtasks", [])
        for i, subtask in enumerate(subtasks):
            text = subtask.get("text", "")
            done = subtask.get("done", False)
            prefix = "✓ " if done else "☐ "
            listbox.insert(tk.END, f"{prefix}{text}")
            if done:
                listbox.itemconfig(i, fg="gray")
        
        # --- Buttons frame ---
        btn_frame = tk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(0, 8))
        
        def add_subtask():
            """Add a new subtask."""
            # Prompt for subtask text
            subtask_dialog = tk.Toplevel(dialog)
            subtask_dialog.title("➕ 添加支线任务")
            subtask_dialog.geometry("400x210")
            subtask_dialog.resizable(False, False)
            subtask_dialog.transient(dialog)
            # Don't use grab_set() to avoid focus issues
            # subtask_dialog.grab_set()  # Commented out to fix input issue
            
            # Center the dialog
            subtask_dialog.update_idletasks()
            w = subtask_dialog.winfo_width()
            h = subtask_dialog.winfo_height()
            x = (subtask_dialog.winfo_screenwidth() // 2) - (w // 2)
            y = (subtask_dialog.winfo_screenheight() // 2) - (h // 2)
            subtask_dialog.geometry(f"{w}x{h}+{x}+{y}")
            
            f = tk.Frame(subtask_dialog, padx=20, pady=20)
            f.pack(fill=tk.BOTH, expand=True)
            
            tk.Label(f, text="支线任务描述:", font=self.font_label).pack(anchor=tk.W, pady=(0, 8))

            text_var = tk.StringVar()
            entry = tk.Entry(f, font=self.font_entry, textvariable=text_var)
            entry.pack(fill=tk.X, pady=(0, 12))
            
            # 标签（逗号分隔）
            tk.Label(f, text=self.T['dlg_tag'], font=self.font_label).pack(anchor=tk.W, pady=(0, 4))
            tags_var = tk.StringVar()
            tags_entry = tk.Entry(f, font=self.font_entry, textvariable=tags_var)
            tags_entry.pack(fill=tk.X, pady=(0, 12))
            # Use after() to ensure focus is set after dialog is fully created
            entry.after(100, lambda: entry.focus_force())
            
            def save_subtask():
                text = text_var.get().strip()
                if not text:
                    messagebox.showwarning("警告", "支线任务描述不能为空！")
                    return
                
                # 校验标签
                tags_text = tags_var.get().strip()
                is_valid, error_msg, tags = self.validate_tags(tags_text)
                if not is_valid:
                    messagebox.showwarning("警告", error_msg)
                    return
                
                # Add subtask
                if "subtasks" not in task:
                    task["subtasks"] = []
                
                new_subtask = {"id": len(task["subtasks"]) + 1, "text": text, "done": False,
                              "tags": tags, "created": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
                task["subtasks"].append(new_subtask)
                
                # Update listbox
                listbox.insert(tk.END, f"☐ {text}")
                
                # Save data
                self.save_data()
                self.refresh_display()
                
                subtask_dialog.destroy()
            
            bf = tk.Frame(f)
            bf.pack(fill=tk.X)
            
            tk.Button(bf, text="✅ 保存", command=save_subtask,
                      bg="#5CB85C", fg="white", font=self.font_button,
                      cursor="hand2", padx=16, pady=4).pack(side=tk.RIGHT, padx=(8, 0))
            
            tk.Button(bf, text="取消", command=subtask_dialog.destroy,
                      font=self.font_button,
                      cursor="hand2", padx=16, pady=4).pack(side=tk.RIGHT)
        
        def toggle_subtask():
            """Toggle subtask done status."""
            selection = listbox.curselection()
            if not selection:
                messagebox.showwarning("警告", "请选择要切换的支线任务！")
                return
            
            index = selection[0]
            if index < len(task.get("subtasks", [])):
                subtask = task["subtasks"][index]
                subtask["done"] = not subtask.get("done", False)
                
                # Update listbox
                text = subtask.get("text", "")
                done = subtask.get("done", False)
                prefix = "✓ " if done else "☐ "
                listbox.delete(index)
                listbox.insert(index, f"{prefix}{text}")
                if done:
                    listbox.itemconfig(index, fg="gray")
                else:
                    listbox.itemconfig(index, fg="black")
                
                # Save data
                self.save_data()
                self.refresh_display()
        
        def delete_subtask():
            """Delete selected subtask."""
            selection = listbox.curselection()
            if not selection:
                messagebox.showwarning("警告", "请选择要删除的支线任务！")
                return
            
            index = selection[0]
            if index < len(task.get("subtasks", [])):
                # Confirm deletion
                if messagebox.askyesno("确认", "确定要删除这个支线任务吗？"):
                    task["subtasks"].pop(index)
                    
                    # Update listbox
                    listbox.delete(index)
                    
                    # Save data
                    self.save_data()
                    self.refresh_display()
        
        tk.Button(btn_frame, text="➕ 添加支线任务", command=add_subtask,
                  bg="#5CB85C", fg="white", font=self.font_button,
                  cursor="hand2", padx=12, pady=4).pack(side=tk.LEFT, padx=(0, 8))
        
        tk.Button(btn_frame, text="✓ 切换完成", command=toggle_subtask,
                  bg="#F0AD4E", fg="white", font=self.font_button,
                  cursor="hand2", padx=12, pady=4).pack(side=tk.LEFT, padx=(0, 8))
        
        tk.Button(btn_frame, text="🗑 删除支线任务", command=delete_subtask,
                  bg="#D9534F", fg="white", font=self.font_button,
                  cursor="hand2", padx=12, pady=4).pack(side=tk.LEFT)
        
        # --- Close button ---
        tk.Button(main_frame, text="关闭", command=dialog.destroy,
                  font=self.font_button,
                  cursor="hand2", padx=20, pady=4).pack(pady=(8, 0))
    
    def show_add_task_dialog(self):
        """Show dialog to add a new task with all fields."""
        # ── Build add task dialog ──
        dialog = tk.Toplevel(self.root)
        dialog.title(self.T['dlg_title_receive'])
        dialog.geometry("520x450")  # Increased height to show all fields
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center the dialog
        dialog.update_idletasks()
        width = dialog.winfo_width()
        height = dialog.winfo_height()
        x = (dialog.winfo_screenwidth() // 2) - (width // 2)
        y = (dialog.winfo_screenheight() // 2) - (height // 2)
        dialog.geometry(f"{width}x{height}+{x}+{y}")
        
        main_frame = tk.Frame(dialog, padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # --- 📝 冒险描述 ---
        tk.Label(main_frame, text=self.T['lbl_desc_edit'], font=self.font_label).pack(anchor=tk.W, pady=(0, 4))
        text_var = tk.StringVar()
        text_entry = tk.Entry(main_frame, font=self.font_entry, textvariable=text_var)
        text_entry.pack(fill=tk.X, pady=(0, 14))
        text_entry.focus_set()
        
        # --- 🗺️ 冒险类型 & ⚡ 紧急等级 ---
        row_frame = tk.Frame(main_frame)
        row_frame.pack(fill=tk.X, pady=(0, 14))
        
        g1 = tk.Frame(row_frame)
        g1.pack(side=tk.LEFT, padx=(0, 20))
        tk.Label(g1, text=self.T['lbl_type'], font=self.font_label).pack(anchor=tk.W, pady=(0, 4))
        category_var = tk.StringVar(value=self.category_options[0])
        category_combo = ttk.Combobox(g1, textvariable=category_var,
                                      values=self.category_options,
                                      state='readonly', width=16, font=self.font_entry)
        category_combo.pack()
        
        g2 = tk.Frame(row_frame)
        g2.pack(side=tk.LEFT)
        tk.Label(g2, text=self.T['lbl_priority'], font=self.font_label).pack(anchor=tk.W, pady=(0, 4))
        priority_var = tk.StringVar(value=self.priority_options[0])
        priority_combo = ttk.Combobox(g2, textvariable=priority_var,
                                      values=self.priority_options,
                                      state='readonly', width=16, font=self.font_entry)
        priority_combo.pack()
        
        # --- ⚔️ 冒险者 ---
        tk.Label(main_frame, text=self.T['lbl_owner'], font=self.font_label).pack(anchor=tk.W, pady=(0, 4))
        person_var = tk.StringVar(value="韦程")
        person_entry = tk.Entry(main_frame, font=self.font_entry, textvariable=person_var)
        person_entry.pack(fill=tk.X, pady=(0, 14))
        
        # --- 🏷️ 标签 ---
        tk.Label(main_frame, text=self.T['dlg_tag2'], font=self.font_label).pack(anchor=tk.W, pady=(0, 4))
        tags_var = tk.StringVar()
        tags_entry = tk.Entry(main_frame, font=self.font_entry, textvariable=tags_var)
        tags_entry.pack(fill=tk.X, pady=(0, 18))
        
        # --- Buttons ---
        btn_frame = tk.Frame(main_frame)
        btn_frame.pack(fill=tk.X)
        
        def save_new_task():
            """Save the new task"""
            task_text = text_var.get().strip()
            if not task_text:
                messagebox.showwarning("警告", self.T['warn_enter_desc'])
                return
            
            # Process tags
            tags_text = tags_var.get().strip()
            tags = [tag.strip() for tag in tags_text.split(',') if tag.strip()] if tags_text else []
            
            # Calculate new task ID
            all_tasks = self.data["tasks"] + self.data["completed"]
            all_ids = [t["id"] for t in all_tasks if "id" in t]
            new_id = max(all_ids) + 1 if all_ids else 1
            
            # Get person
            person_val = person_var.get().strip()
            if not person_val:
                person_val = "韦程"
            
            # Create new task
            task = {
                "id": new_id,
                "text": task_text,
                "priority": priority_var.get(),
                "category": category_var.get(),
                "person": person_val,
                "tags": tags,
                "created": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            
            self.data["tasks"].append(task)
            self.save_data()
            self.refresh_display()
            self.set_status(self.T['status_task_received'].format(task_text))
            dialog.destroy()
        
        tk.Button(btn_frame, text=self.T['btn_save_new'], command=save_new_task,
                  bg=THEME['colors']['primary'], fg="white", font=self.font_button,
                  cursor="hand2", padx=20, pady=4).pack(side=tk.RIGHT, padx=(10, 0))
        
        tk.Button(btn_frame, text="取消", command=dialog.destroy,
                  font=self.font_button,
                  cursor="hand2", padx=20, pady=4).pack(side=tk.RIGHT)
        self._apply_emoji_images(dialog)
    
    def refresh_display(self):
        """Refresh the display"""
        self.refresh_pending()
        self.refresh_completed()
        self.refresh_stats()
        # 刷新后立即更新成长系统显示（每周智慧系统面板）
        self.update_game_display()
    
    def refresh_pending(self):
        """Refresh pending tasks display - sorted by selected criteria"""
        self.pending_tree.delete(*self.pending_tree.get_children())
        
        # Get sorting criteria
        criteria = self.sort_criteria.get()
        
        # Sort tasks based on criteria
        if criteria == self.T['sort_default']:
            # 优先级排序：红色警报 > 修炼升级 > 临时密令 > 佛系摸鱼
            priority_order = dict(zip(self.priority_options, [1, 2, 3, 4]))
            sorted_tasks = sorted(self.data["tasks"], 
                               key=lambda t: priority_order.get(t.get("priority", ""), 999))
        elif criteria == self.T['sort_priority']:
            # 按紧急程度排序（同默认顺序）
            priority_order = dict(zip(self.priority_options, [1, 2, 3, 4]))
            sorted_tasks = sorted(self.data["tasks"], 
                               key=lambda t: priority_order.get(t.get("priority", ""), 999))
        elif criteria == self.T['sort_created']:
            # 按创建时间排序（新到旧）
            sorted_tasks = sorted(self.data["tasks"], 
                               key=lambda t: t.get("created", ""), reverse=True)
        elif criteria == self.T['sort_desc']:
            # 按任务描述排序（A-Z）
            sorted_tasks = sorted(self.data["tasks"], 
                               key=lambda t: t.get("text", "").lower())
        else:
            # 默认：按ID排序
            sorted_tasks = self.data["tasks"]
        
        # Filter by search query
        query = self.search_query.get().strip().lower()
        if query:
            filtered_tasks = []
            for task in sorted_tasks:
                # Search in: description, person, category, priority, tags
                text = task["text"].lower()
                person = self.format_person(task.get("person", "")).lower()
                category = task.get("category", "").lower()
                priority = task.get("priority", "").lower()
                tags = ", ".join(task.get("tags", [])).lower()
                
                if query in text or query in person or query in category or query in priority or query in tags:
                    filtered_tasks.append(task)
            sorted_tasks = filtered_tasks
        
        for display_id, task in enumerate(sorted_tasks, 1):
            # Format tags for display
            tags = task.get("tags", [])
            tags_str = ", ".join(tags) if tags else ""
            
            # Calculate progress
            subtasks = task.get("subtasks", [])
            total = len(subtasks)
            done = sum(1 for st in subtasks if st.get("done", False))
            progress = f"{done}/{total}" if total > 0 else "─"
            
            # Insert main task row
            item = self.pending_tree.insert('', tk.END, iid=str(task["id"]), values=(
                '☐',  # 复选框列 - 未选中
                display_id,
                task["text"],
                task.get("category", ""),
                task.get("priority", ""),
                self.format_person(task.get("person", "")),
                tags_str,  # Display tags
                progress,  # Display progress
                task["created"]
            ))
            
            priority = task.get("priority", "")
            if priority in self.priority_colors:
                self.pending_tree.item(item, tags=('main', priority,))
            
            # Insert subtask rows directly below main task
            for subtask_idx, subtask in enumerate(subtasks, 1):
                subtask_id = f"{task['id']}_subtask_{subtask_idx}"
                subtask_display_id = f"{display_id}.{subtask_idx}"
                subtask_text = subtask.get("text", "")
                subtask_done = subtask.get("done", False)
                
                # Apply strikethrough for completed subtasks（层级由 tree 缩进承担，不再额外加空格）
                if subtask_done:
                    subtask_display_text = self.strikethrough(subtask_text)
                else:
                    subtask_display_text = subtask_text
                
                # 支线标签与领取时间
                subtask_tags = subtask.get("tags", [])
                subtask_tags_str = ", ".join(subtask_tags) if subtask_tags else ""
                subtask_created = subtask.get("created", "")
                
                # Add subtask row（挂到主线 item 下，形成树形层级）
                subtask_item = self.pending_tree.insert(item, tk.END, iid=subtask_id, values=(
                    '☑️' if subtask_done else '☐',  # Subtask checkbox
                    subtask_display_id,  # Subtask ID (e.g., "1.1", "1.2")
                    subtask_display_text,  # Indented subtask text (with strikethrough if done)
                    "",  # No category for subtasks
                    "",  # No priority for subtasks
                    "",  # No person for subtasks
                    subtask_tags_str,  # 支线标签
                    "✓" if subtask_done else "─",  # Subtask status
                    subtask_created  # 领取时间
                ))
                
                # Apply gray background for completed subtasks
                if subtask_done:
                    self.pending_tree.item(subtask_item, tags=('subtask_done',))
                else:
                    self.pending_tree.item(subtask_item, tags=('subtask',))

            # 树形层级默认全部展开（支线默认可见，由用户点三角手动收起；无支线主线为叶子不影响）
            self.pending_tree.item(item, open=True)
        
        for priority, color in self.priority_colors.items():
            self.pending_tree.tag_configure(priority, background=color)
        
        # Configure subtask tags
        self.pending_tree.tag_configure('subtask', background='#F5F5F5')  # Light gray for subtasks
        self.pending_tree.tag_configure('subtask_done', background='#E8E8E8', foreground='gray')  # Darker gray for completed subtasks
        
    def refresh_completed(self):
        """Refresh completed tasks display"""
        self.completed_tree.delete(*self.completed_tree.get_children())
        
        dates = set()
        for task in self.data["completed"]:
            if "completed" in task:
                date = task["completed"].split()[0]
                dates.add(date)
        
        dates = sorted(list(dates), reverse=True)
        self.date_filter['values'] = ['全部'] + dates
        if not self.date_filter.get():
            self.date_filter.set('全部')
        
        self.filter_completed()
    
    def filter_completed(self):
        """Filter completed tasks by date, sorted by completion time (newest first)"""
        self.completed_tree.delete(*self.completed_tree.get_children())
        
        selected_date = self.date_filter.get()
        
        # 过滤并排序
        filtered_tasks = []
        for task in self.data["completed"]:
            if "completed" in task:
                task_date = task["completed"].split()[0]
                if selected_date == '全部' or task_date == selected_date:
                    filtered_tasks.append(task)
        
        # Filter by search query
        query = self.search_query.get().strip().lower()
        if query:
            filtered_tasks2 = []
            for task in filtered_tasks:
                # Search in: description, person, category, priority, tags
                text = task["text"].lower()
                person = self.format_person(task.get("person", "")).lower()
                category = task.get("category", "").lower()
                priority = task.get("priority", "").lower()
                tags = ", ".join(task.get("tags", [])).lower()
                
                if query in text or query in person or query in category or query in priority or query in tags:
                    filtered_tasks2.append(task)
            filtered_tasks = filtered_tasks2
        
        # Sort based on criteria
        criteria = self.sort_criteria.get()
        if criteria == self.T['sort_default'] or criteria == self.T['sort_completed']:
            # 按完成时间倒序排列(最新的在前面)
            filtered_tasks.sort(key=lambda t: t.get("completed", ""), reverse=True)
        elif criteria == self.T['sort_priority']:
            # 按紧急程度排序
            priority_order = dict(zip(self.priority_options, [1, 2, 3, 4]))
            filtered_tasks.sort(key=lambda t: priority_order.get(t.get("priority", ""), 999))
        elif criteria == self.T['sort_created']:
            # 按创建时间排序（新到旧）
            filtered_tasks.sort(key=lambda t: t.get("created", ""), reverse=True)
        elif criteria == self.T['sort_desc']:
            # 按任务描述排序（A-Z）
            filtered_tasks.sort(key=lambda t: t.get("text", "").lower())
        else:
            # 默认：按完成时间倒序
            filtered_tasks.sort(key=lambda t: t.get("completed", ""), reverse=True)
        
        # 获取今天日期
        today = datetime.now().strftime("%Y-%m-%d")
        
        for display_id, task in enumerate(filtered_tasks, 1):
            duration = self.calculate_duration(task.get("created", ""), task["completed"])
            # 应用删除线效果
            task_text = self.strikethrough(task["text"])
            
            # Format tags for display
            tags = task.get("tags", [])
            tags_str = ", ".join(tags) if tags else ""
            
            # Calculate progress
            subtasks = task.get("subtasks", [])
            total = len(subtasks)
            done = sum(1 for st in subtasks if st.get("done", False))
            progress = f"{done}/{total}" if total > 0 else "─"
            
            # Insert main task row
            item = self.completed_tree.insert('', tk.END, iid=str(task["id"]), values=(
                '☑️',  # 复选框列 - 已选中
                display_id,
                task_text,
                task.get("category", ""),
                task.get("priority", ""),
                self.format_person(task.get("person", "")),
                tags_str,  # Display tags
                progress,  # Display progress
                task["completed"],
                duration
            ))
            
            # 判断是否是当天完成,设置不同背景色
            task_date = task["completed"].split()[0]
            if task_date == today:
                # 当天完成:深灰色
                self.completed_tree.item(item, tags=('completed_today',))
            else:
                # 超过1天:浅灰色
                self.completed_tree.item(item, tags=('completed_old',))
        
    # ──────────────────────────────────────────────
    #  STATISTICS — Table-based, week-aware
    # ──────────────────────────────────────────────
    def _date_to_week(self, date_str):
        """Parse a date string (YYYY-MM-DD ...) and return (iso_year, iso_week)."""
        try:
            d = datetime.strptime(date_str.split()[0], "%Y-%m-%d")
            iso = d.isocalendar()
            return (iso[0], iso[1])
        except (ValueError, IndexError):
            return None

    def _build_week_options(self):
        """Build week dropdown options from all task dates."""
        weeks_set = set()
        for t in self.data["tasks"]:
            w = self._date_to_week(t.get("created", ""))
            if w:
                weeks_set.add(w)
        for t in self.data["completed"]:
            w = self._date_to_week(t.get("completed", ""))
            if w:
                weeks_set.add(w)
        # 当前周（即使本周暂无任务也确保出现在下拉首位）
        now = datetime.now()
        iso = now.isocalendar()
        cur_key = (iso[0], iso[1])
        weeks_set.add(cur_key)
        # 当前周置顶，其余周按时间倒序（最近的在前）
        others = sorted((w for w in weeks_set if w != cur_key), reverse=True)
        self._all_weeks = [cur_key] + others
        cur_label = f"{cur_key[0]}年第{cur_key[1]}周"
        options = [cur_label] + [f"{y}年第{w}周" for y, w in others] + ["全部"]
        self.week_filter_combo['values'] = options
        if not self.week_filter_var.get() or self.week_filter_var.get() not in options:
            self.week_filter_var.set(options[0] if options else "全部")

    def _filter_by_week(self, tasks, date_field):
        """Filter a list of tasks by the currently selected week.
        Returns filtered list, or original list if selected week is '全部'.
        """
        selected = self.week_filter_var.get()
        if selected == "全部":
            return tasks
        m = re.match(r"(\d+)年.*?(\d+)周", selected)
        if not m:
            return tasks
        year_filter, week_filter = int(m.group(1)), int(m.group(2))
        result = []
        for t in tasks:
            w = self._date_to_week(t.get(date_field, ""))
            if w and w[0] == year_filter and w[1] == week_filter:
                result.append(t)
        return result

    def _update_clock(self):
        """刷新战绩页「选择周」右侧的实时时钟（描边幽灵牌），每秒一次"""
        now = datetime.now()
        weekday_cn = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日'][now.weekday()]
        text = f"{now.year}年{now.month:02d}月{now.day:02d}日 {weekday_cn}  {now.hour:02d}：{now.minute:02d}：{now.second:02d}"
        if getattr(self, 'clock_text', None):
            self.clock_text.config(text=text)
        if getattr(self, 'clock_canvas', None):
            self._draw_clock(self.clock_canvas, now)
        self.root.after(1000, self._update_clock)

    def _draw_clock(self, canvas, now):
        """在 16x16 Canvas 上绘制蓝色线条时钟（圆 + 时分秒针），颜色与描边同色"""
        canvas.delete("all")
        cx, cy, r = 8, 8, 6
        canvas.create_oval(cx - r, cy - r, cx + r, cy + r, outline=THEME['colors']['primary'], width=1.5)
        # 指针角度：12 点方向为 0，顺时针。Tk 的 y 轴向下，故 y 取负
        sec = now.second / 60.0
        minute = (now.minute + now.second / 60.0) / 60.0
        hour = ((now.hour % 12) + now.minute / 60.0) / 12.0
        # 秒针（蓝，细）
        sx = cx + (r - 1) * math.sin(sec * 2 * math.pi)
        sy = cy - (r - 1) * math.cos(sec * 2 * math.pi)
        canvas.create_line(cx, cy, sx, sy, fill=THEME['colors']['primary'], width=1)
        # 分针（深蓝，中）
        mx = cx + (r - 2) * math.sin(minute * 2 * math.pi)
        my = cy - (r - 2) * math.cos(minute * 2 * math.pi)
        canvas.create_line(cx, cy, mx, my, fill=THEME['colors']['title_bar'], width=1.2)
        # 时针（深蓝，粗）
        hx = cx + (r - 3) * math.sin(hour * 2 * math.pi)
        hy = cy - (r - 3) * math.cos(hour * 2 * math.pi)
        canvas.create_line(cx, cy, hx, hy, fill=THEME['colors']['title_bar'], width=1.4)

    def _on_resize(self, event):
        """窗口尺寸变化回调（V1.6.18）：仅当『从最大化→还原』这一跳变时，
        强制锁回首开基准尺寸 self.window_home，保证还原后与小窗口首开一致；
        手动拖拽缩放不受影响。"""
        if event.widget is not self.root:
            return  # 忽略子控件尺寸变化冒泡上来的 Configure 事件
        st = self.root.wm_state()
        if self._prev_state == "zoomed" and st == "normal":
            self.root.geometry(self.window_home)
        self._prev_state = st

    def refresh_stats(self):
        """Refresh statistics tab data, filtered by selected week"""
        self._build_week_options()
        
        # Filter by week
        filtered_tasks = self._filter_by_week(self.data["tasks"], "created")
        filtered_completed = self._filter_by_week(self.data["completed"], "completed")
        
        total_pending = len(filtered_tasks)
        total_completed = len(filtered_completed)
        total_all = total_pending + total_completed
        
        self.stat_pending.set(str(total_pending))
        self.stat_completed.set(str(total_completed))
        if total_all > 0:
            rate = total_completed / total_all * 100
            self.stat_rate.set(f"{rate:.1f}%")
        else:
            self.stat_rate.set("暂无数据")
        
        # ── By priority ──
        self.stat_priority_tree.delete(*self.stat_priority_tree.get_children())
        for pri in self.priority_options:
            p = sum(1 for t in filtered_tasks if t.get("priority") == pri)
            c = sum(1 for t in filtered_completed if t.get("priority") == pri)
            total = p + c
            if p > 0 or c > 0:
                self.stat_priority_tree.insert('', tk.END, values=(pri, p, c, total))
        
        # ── By category ──
        self.stat_category_tree.delete(*self.stat_category_tree.get_children())
        for cat in self.category_options:
            p = sum(1 for t in filtered_tasks if t.get("category") == cat)
            c = sum(1 for t in filtered_completed if t.get("category") == cat)
            total = p + c
            if p > 0 or c > 0:
                self.stat_category_tree.insert('', tk.END, values=(cat, p, c, total))
        
        # ── By date ──
        self._stat_date_counts = {}
        for t in filtered_completed:
            d = t.get("completed", "").split()[0]
            self._stat_date_counts[d] = self._stat_date_counts.get(d, 0) + 1
        # 填充下拉选项
        dates = sorted(self._stat_date_counts.keys(), reverse=True)
        self.stat_date_filter['values'] = ['全部'] + dates
        if not self.stat_date_filter.get() or self.stat_date_filter.get() not in (['全部'] + dates):
            self.stat_date_filter.set('全部')
        # 按选中日期刷新表格
        self._refresh_stat_date()
        
    def _refresh_stat_date(self):
        """按下拉选中的日期刷新『按日期统计（已通关）』表格"""
        self.stat_date_tree.delete(*self.stat_date_tree.get_children())
        if not hasattr(self, '_stat_date_counts'):
            return
        sel = self.stat_date_filter.get() if self.stat_date_filter.get() else '全部'
        if sel == '全部':
            for date in sorted(self._stat_date_counts.keys(), reverse=True):
                self.stat_date_tree.insert('', tk.END,
                                           values=(date, self._stat_date_counts[date]))
        else:
            self.stat_date_tree.insert('', tk.END,
                                       values=(sel, self._stat_date_counts.get(sel, 0)))

def main():
    # Check if another instance is already running
    if is_already_running():
        root = tk.Tk()
        root.withdraw()
        messagebox.showwarning(
            "提示",
            "TodoList 已经在运行中！\n\n请勿同时打开多个窗口，以免造成数据混乱。"
        )
        root.destroy()
        sys.exit(0)
    
    root = tk.Tk()
    app = TodoApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
